#  Call_Empty = fix_fastrow
#         if fix_column > fix_endrow:
#            Row_insart = fix_column - fix_endrow
           
#            for i in range(Row_insart):
                
#                 ws2.insert_rows(fix_maxrow)
#                 ws2['A'+ str(fix_maxrow)].number_format = u'm\月d\日'
#                 for j in ws2["A1":"J1000"]:
#                     for l in j:
#                         l.border = border
#                 for j in ws2["L1":"M1000"]:
#                     for l in j:
#                         l.border = border
#                 fix_endrow += 1
#                 wb2.save(failpass2)
#         elif fix_column < fix_endrow:
#             Row_delete = fix_endrow - fix_column
#             ws2.delete_rows(idx=fix_fastrow, amount=Row_delete)
#             wb2.save(failpass2)

def save_contens(comfirm):
    global Call_Empty
    global All_contents
    global fix_endrow
    global All_contents
    border = Border(top=side, bottom=side, left=side, right=side)
    wb2 = openpyxl.load_workbook(failpass2, keep_vba=True)
    ws2 = wb2.worksheets[0]
    if fix_mode == 1:
        Call_Empty = fix_fastrow
        if fix_column > fix_endrow:
           Row_insart = fix_column - fix_endrow
           
           for i in range(Row_insart):
                
                ws2.insert_rows(fix_maxrow)
                ws2['A'+ str(fix_maxrow)].number_format = u'm\月d\日'
                for j in ws2["A1":"J1000"]:
                    for l in j:
                        l.border = border
                for j in ws2["L1":"M1000"]:
                    for l in j:
                        l.border = border
                fix_endrow += 1
        elif fix_column < fix_endrow:
            Row_delete = fix_endrow - fix_column
            ws2.delete_rows(idx=fix_fastrow, amount=Row_delete)
            
    for n in range(len(comfirm)):
        b = 1
        for row_a in comfirm[n]:
            ws2.cell(row=Call_Empty,column=b).value = row_a
            b+=1
        Call_Empty +=1
    All_contents = []
    try:
        wb2.save(failpass2)
        messagebox.showwarning("確認","登録しました。")
        allclear()
    except:
        messagebox.showwarning("確認","登録ファイルが開かれています。")
        
    

def search():
    searchbox_listbox.delete(0,END)
    searchbox_entry_get = str(searchbox_entry.get())
    for i in mat_number_name_str:
        if searchbox_entry_get in i:
            searchbox_listbox.insert(tk.END,i)

def search_clear():
    searchbox_listbox.delete(0,END)
    searchbox_entry.delete(0,tk.END)



def search_insert():
    global comfirm_mat_num_1
    global comfirm_mat_num_2
    global comfirm_mat_num_3
    global comfirm_mat_num_4
    global comfirm_mat_num_5
    global comfirm_mat_num_6
    global comfirm_mat_num_7
    global comfirm_mat_num_8
    mat_index = searchbox_listbox.curselection()[0]
    
    if checkbutton_var_1.get() == "1":
        f = comfirm_mat_num_1 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_1.current(f)

    if checkbutton_var_2.get() == "1":
        f = comfirm_mat_num_2 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_2.current(f)

    if checkbutton_var_3.get() == "1":
        f = comfirm_mat_num_3 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_3.current(f)

    if checkbutton_var_4.get() == "1":
        f = comfirm_mat_num_4 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_4.current(f)

    if checkbutton_var_5.get() == "1":
        f = comfirm_mat_num_5 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_5.current(f)

    if checkbutton_var_6.get() == "1":
        f = comfirm_mat_num_6 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_6.current(f)

    if checkbutton_var_7.get() == "1":
        f = comfirm_mat_num_7 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_7.current(f)

    if checkbutton_var_8.get() == "1":
        f = comfirm_mat_num_8 = searchbox_listbox.get(mat_index)[:7]
        f = mat_number.index(str(f))
        mat_num_combobox_8.current(f)
    
def cell_check(ws_v, ws_d):
    cell_check_value = loadingpass_ws['D2'].value
    if cell_check_value == None:
        cell_check_mode = 0
    else:
        cell_check_mode = cell_check_value
    
    if int(cell_check_mode) == 1:
        j = len(ws_v)
        if j > 100:
            i = ws_d.max_row + 1
        else:
            i = 0
            for cell_i in ws_v:
                i += 1
                if cell_i.value == None:
                    break
    elif int(cell_check_mode) == 0:
        i = 0
        for cell_i in ws_v:
                i += 1
                if cell_i.value == None:
                    break
        if cell_i.value != None:
            i = 0
            for cell_i in ws_v:
                    i += 1
                    if cell_i.value == None:
                        break


    return i

def cell_fix():
    global fix_fastrow
    global fix_endrow
    global fix_maxrow
    global fix_mode
    global k
    global comfirm_day_1
    global comfirm_day_2
    global comfirm_day_3
    global comfirm_day_4
    global comfirm_day_5
    global comfirm_day_6
    global comfirm_day_7
    global comfirm_day_8
    global comfirm_mat_num_1
    global comfirm_mat_num_2
    global comfirm_mat_num_3
    global comfirm_mat_num_4
    global comfirm_mat_num_5
    global comfirm_mat_num_6
    global comfirm_mat_num_7
    global comfirm_mat_num_8    
    global comfirm_warkclass_num_1
    global comfirm_warkclass_num_2
    global comfirm_warkclass_num_3
    global comfirm_warkclass_num_4
    global comfirm_warkclass_num_5
    global comfirm_warkclass_num_6
    global comfirm_warkclass_num_7
    global comfirm_warkclass_num_8
    global comfirm_wark_num_1
    global comfirm_wark_num_2
    global comfirm_wark_num_3
    global comfirm_wark_num_4
    global comfirm_wark_num_5
    global comfirm_wark_num_6
    global comfirm_wark_num_7
    global start_time_h1
    global start_time_h2
    global start_time_h3
    global start_time_h4
    global start_time_h5
    global start_time_h6
    global start_time_h7
    global start_time_h8
    global start_time_m1
    global start_time_m2
    global start_time_m3
    global start_time_m4
    global start_time_m5
    global start_time_m6
    global start_time_m7
    global start_time_m8
    global end_time_h1
    global end_time_h2
    global end_time_h3
    global end_time_h4
    global end_time_h5
    global end_time_h6
    global end_time_h7
    global end_time_h8
    global end_time_m1
    global end_time_m2
    global end_time_m3
    global end_time_m4
    global end_time_m5
    global end_time_m6
    global end_time_m7
    global end_time_m8
    global comfirm_start_time_h_1
    global comfirm_start_time_h_2
    global comfirm_start_time_h_3
    global comfirm_start_time_h_4
    global comfirm_start_time_h_5
    global comfirm_start_time_h_6
    global comfirm_start_time_h_7
    global comfirm_start_time_h_8
    global comfirm_start_time_m_1
    global comfirm_start_time_m_2
    global comfirm_start_time_m_3
    global comfirm_start_time_m_4
    global comfirm_start_time_m_5
    global comfirm_start_time_m_6
    global comfirm_start_time_m_7
    global comfirm_start_time_m_8
    global comfirm_end_time_h_1
    global comfirm_end_time_h_2
    global comfirm_end_time_h_3
    global comfirm_end_time_h_4
    global comfirm_end_time_h_5
    global comfirm_end_time_h_6
    global comfirm_end_time_h_7
    global comfirm_end_time_h_8
    global comfirm_end_time_m_1
    global comfirm_end_time_m_2
    global comfirm_end_time_m_3
    global comfirm_end_time_m_4
    global comfirm_end_time_m_5
    global comfirm_end_time_m_6
    global comfirm_end_time_m_7
    global comfirm_end_time_m_8
    
    def_wb2 = openpyxl.load_workbook(failpass2, keep_vba=True)
    def_ws2 = def_wb2.worksheets[0]
    if "comfirm_name" in globals() and "comfirm_month_1" in globals() and "comfirm_day_1" in globals():
        
        _name_ =  comfirm_name
        _date_ = datetime.datetime(int(ny),int(comfirm_month_1),int(comfirm_day_1),00,00)
        i = 0
        def_wb2 = openpyxl.load_workbook(failpass2, keep_vba=True)
        def_ws2 = def_wb2.worksheets[0]

        for (cell_i,cell_j) in zip(def_ws2['B'],def_ws2['A']):
            i += 1
            if cell_i.value == _name_ and cell_j.value == _date_:
                fix_mode = 1
                try:
                    k
                except NameError:
                    k = i
                cells_A = 'A' + str(i + 1)
                cells_B = 'B' + str(i + 1)
                if not def_ws2[cells_B].value == _name_ and def_ws2[cells_A].value == _date_ or not def_ws2[cells_A].value == _date_ or def_ws2[cells_B].value == None:
                    fix_fastrow = k
                    fix_endrow = i
                    fix_maxrow = i
                    
                    fix_endrow = fix_endrow - fix_fastrow + 1
                    break
            elif cell_i.value == None:
                try:
                    fix_fastrow = k
                except NameError:
                        messagebox.showwarning("確認","該当するデータは見つかりませんでした。")
                fix_endrow = i
                fix_endrow = fix_endrow - fix_fastrow + 1
                break

        if "fix_fastrow" in globals() and "fix_endrow" in globals():
            
            i = i - k
            i += 1
            if i >= 1:
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_1.current(m)
                d = comfirm_day_1 = f.day
                d = day_list.index(str(d))
                today_day_combobox_1.current(d)
                
                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_1 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_1.current(f)
                except ValueError:
                    pass
                try:    
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_1 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_1.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（1行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_1 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_1.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h1 = comfirm_start_time_h_1 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_1.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m1 = comfirm_start_time_m_1 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_1.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h1 = comfirm_end_time_h_1 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_1.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m1 = comfirm_end_time_m_1 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_1.current(f)
                
            if i >= 2:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_2.current(m)
                d = comfirm_day_2 = f.day
                d = day_list.index(str(d))
                today_day_combobox_2.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_2 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_2.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_2 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_2.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（2行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_2 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_2.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h2 = comfirm_start_time_h_2 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_2.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m2 = comfirm_start_time_m_2 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_2.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h2 = comfirm_end_time_h_2 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_2.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m2 = comfirm_end_time_m_2 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_2.current(f)

            if i >= 3:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_3.current(m)
                d = comfirm_day_3 = f.day
                d = day_list.index(str(d))
                today_day_combobox_3.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_3 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_3.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_3 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_3.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（3行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_3 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_3.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h3 = comfirm_start_time_h_3 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_3.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m3 = comfirm_start_time_m_3 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_3.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h3 = comfirm_end_time_h_3 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_3.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m3 = comfirm_end_time_m_3 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_3.current(f)

            if i >= 4:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_4.current(m)
                d = comfirm_day_4 = f.day
                d = day_list.index(str(d))
                today_day_combobox_4.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_4 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_4.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_4 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_4.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（4行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_4 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_4.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h4 = comfirm_start_time_h_4 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_4.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m4 = comfirm_start_time_m_4 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_4.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h4 = comfirm_end_time_h_4 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_4.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m4 = comfirm_end_time_m_4 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_4.current(f)

            if i >= 5:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_5.current(m)
                d = comfirm_day_5 = f.day
                d = day_list.index(str(d))
                today_day_combobox_5.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_5 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_5.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_5 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_5.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（5行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_5 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_5.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h5 = comfirm_start_time_h_5 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_5.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m5 = comfirm_start_time_m_5 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_5.current(f)
            
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h5 = comfirm_end_time_h_5 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_5.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m5 = comfirm_end_time_m_5 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_5.current(f)
            
            if i >= 6:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_6.current(m)
                d = comfirm_day_6 = f.day
                d = day_list.index(str(d))
                today_day_combobox_6.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_6 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_6.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_6 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_6.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（6行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_6 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_6.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h6 = comfirm_start_time_h_6 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_6.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m6 = comfirm_start_time_m_6 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_6.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h6 = comfirm_end_time_h_6 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_6.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m6 = comfirm_end_time_m_6 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_6.current(f)

            if i >= 7:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_7.current(m)
                d = comfirm_day_7 = f.day
                d = day_list.index(str(d))
                today_day_combobox_7.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_7 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_7.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_7 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_7.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（7行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_7 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_7.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h7 = comfirm_start_time_h_7 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_7.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m7 = comfirm_start_time_m_7 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_7.current(f)
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h7 = comfirm_end_time_h_7 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_7.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m7 = comfirm_end_time_m_7 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_7.current(f)

            if i >= 8:
                k += 1
                cells_A = 'A' + str(k)
                f = def_ws2[cells_A].value
                m = f.month
                m = month_list.index(str(m))
                today_month_combobox_8.current(m)
                d = comfirm_day_8 = f.day
                d = day_list.index(str(d))
                today_day_combobox_8.current(d)

                try:
                    cells_C = 'C' + str(k)
                    f =  comfirm_mat_num_8 = def_ws2[cells_C].value
                    f = mat_number.index(str(f))
                    mat_num_combobox_8.current(f)
                except ValueError:
                    pass
                try:
                    cells_D = 'D' + str(k)
                    f = comfirm_warkclass_num_8 = def_ws2[cells_D].value
                    f = warkclass_number.index(str(f))
                    warkclass_num_combobox_8.current(f)
                except ValueError:
                    messagebox.showwarning("確認","作業区分の詳細を入力していたため作業区分は削除されました、再度入力してください。（8行目）")

                cells_E = 'E' + str(k)
                f = comfirm_wark_num_8 = def_ws2[cells_E].value
                f = wark_number.index(str(f))
                wark_num_combobox_8.current(f)
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_h8 = comfirm_start_time_h_8 = f.hour
                f = time_h_list.index(str(f))
                time_h_num_combobox_8.current(f) 
                
                cells_F = 'F' + str(k)
                f = def_ws2[cells_F].value
                f = start_time_m8 = comfirm_start_time_m_8 = f.minute
                f = time_m_list.index(str(f))
                time_m_num_combobox_8.current(f)
            
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_h8 = comfirm_end_time_h_8 = f.hour
                f = time_h_list.index(str(f))
                endtime_h_num_combobox_8.current(f) 
                
                cells_G = 'G' + str(k)
                f = def_ws2[cells_G].value
                f = end_time_m8 = comfirm_end_time_m_8 = f.minute
                f = endtime_m_list.index(str(f))
                endtime_m_num_combobox_8.current(f)
        
        if 'end_time_m1' in globals(): 
            end_number_change_m1(k) 
        if 'end_time_m2' in globals(): 
            end_number_change_m2(k) 
        if 'end_time_m3' in globals(): 
            end_number_change_m3(k) 
        if 'end_time_m4' in globals(): 
            end_number_change_m4(k) 
        if 'end_time_m5' in globals(): 
            end_number_change_m5(k) 
        if 'end_time_m6' in globals(): 
            end_number_change_m6(k) 
        if 'end_time_m7' in globals(): 
            end_number_change_m7(k) 
        if 'end_time_m8' in globals(): 
            end_number_change_m8(k) 
        
        del k

def Calculation():
    if 'end_time_m1' in globals(): 
        end_number_change_m1("k") 
    if 'end_time_m2' in globals(): 
        end_number_change_m2("k") 
    if 'end_time_m3' in globals(): 
        end_number_change_m3("k") 
    if 'end_time_m4' in globals(): 
        end_number_change_m4("k") 
    if 'end_time_m5' in globals(): 
        end_number_change_m5("k") 
    if 'end_time_m6' in globals(): 
        end_number_change_m6("k") 
    if 'end_time_m7' in globals(): 
        end_number_change_m7("k") 
    if 'end_time_m8' in globals(): 
        end_number_change_m8("k") 
    
# def overwark_check():
#     global overwark_check_mode
#     if (overwark_check_combobox['state'] == "noemal"):
#         overwark_check_combobox['state'] = "disabled"
#         overwark_check_mode = 0
#     else:
#         overwark_check_combobox['state'] = "noemal"
#         overwark_check_mode = 1

def one_set():
    global comfirm_end_time_h_1
    global comfirm_end_time_m_1
    global end_time_h1
    global end_time_m1
    comfirm_end_time_h_1 = end_time_h1 = 17
    comfirm_end_time_m_1 = end_time_m1 = 30

    endtime_h_num_combobox_1.current(9)
    endtime_m_num_combobox_1.current(3)
    end_number_change_h1(0)
    end_number_change_m1(0)

def list_state():
    global mode
    
    if break_listbox_1['state'] == "normal":
        break_listbox_1['state'] = "disabled"
        break_listbox_2['state'] = "disabled"
        break_listbox_3['state'] = "disabled"
        break_listbox_4['state'] = "disabled"
        break_listbox_5['state'] = "disabled"
        break_listbox_6['state'] = "disabled"
        break_listbox_7['state'] = "disabled"
        break_listbox_8['state'] = "disabled"
        mode = 0
        
        any_label.place_forget()
    else:
        any_label.place(x=620, y=8)
        break_listbox_1['state'] = "normal"
        break_listbox_2['state'] = "normal"
        break_listbox_3['state'] = "normal"
        break_listbox_4['state'] = "normal"
        break_listbox_5['state'] = "normal"
        break_listbox_6['state'] = "normal"
        break_listbox_7['state'] = "normal"
        break_listbox_8['state'] = "normal"
        mode = 1
        print(break_listbox_1['state'])
        pass

def allclear():

    global comfirm_name
    global mode
    global fix_mode
    global All_contents
    if 'comfirm_name' in globals():
        del comfirm_name
    total_time_min_v.set("合計時間 : "+ "0" + " 分")
    overwark_time_min_v.set("残業時間 : "+ "0" + " 分")
    nightshift_time_min_v.set("夜勤時間 : "+ "0" + " 分")
    ward_var_1.set("名前選択")
    name_num_combobox.set("")
    # overwark_check_combobox.set("")
    any_label.place_forget()
    break_listbox_1['state'] = "disabled"
    break_listbox_2['state'] = "disabled"
    break_listbox_3['state'] = "disabled"
    break_listbox_4['state'] = "disabled"
    break_listbox_5['state'] = "disabled"
    break_listbox_6['state'] = "disabled"
    break_listbox_7['state'] = "disabled"
    break_listbox_8['state'] = "disabled"

    # overwark_check_combobox['state'] = "disabled"
    All_contents = []
    overwark_check_mode = 0
    mode = 0
    nightshift_time = 0
    comfirm_overwark_check = 0
    fix_mode = 0

    search_clear()
    row1clear()
    row2clear()
    row3clear()
    row4clear()
    row5clear()
    row6clear()
    row7clear()
    row8clear()
    
def row1clear():
    global comfirm_mat_num_1
    global comfirm_wark_num_1
    global comfirm_warkclass_num_1
    global comfirm_start_time_h_1
    global comfirm_start_time_m_1
    global comfirm_end_time_h_1
    global comfirm_end_time_m_1
    global comfirm_break_time_1
    global comfirm_nightshift_1
    global comfirm_warkclass_detail_1
    global comfirm_month_1
    global comfirm_day_1
    global total_time_1
    global total_time
    global time_h_num_combobox_1
    global time_m_num_combobox_1
    global start_time_h1
    global start_time_m1
    global end_time_h1
    global end_time_m1
    global today_month_combobox_1
    global today_day_combobox_1

    if 'comfirm_mat_num_1' in globals():
        del comfirm_mat_num_1
    if 'comfirm_wark_num_1' in globals():
        del comfirm_wark_num_1
    if 'comfirm_start_time_h_1' in globals():
        start_time_h1 = comfirm_start_time_h_1 = 8
    if 'comfirm_start_time_m_1' in globals():
        if kanto_mode == 0:
            start_time_m1 = comfirm_start_time_m_1 = 10
        elif kanto_mode == 1:
            start_time_m1 = comfirm_start_time_m_1 = 00    
    if 'comfirm_end_time_h_1' in globals():
        del comfirm_end_time_h_1
    if 'comfirm_end_time_m_1' in globals():
        del comfirm_end_time_m_1
    if 'total_time_1' in globals():
        total_time_1 = 0
    if 'total_time' in globals():
        total_time = 0
        
    if 'comfirm_break_time_1' in globals():
        comfirm_break_time_1 = 0
    if 'comfirm_warkclass_num_1' in globals():
        comfirm_warkclass_num_1 = 0
    if 'end_time_h1' in globals():
        del end_time_h1
    if 'end_time_m1' in globals():
        del end_time_m1
    if 'comfirm_warkclass_detail_1' in globals():
        del comfirm_warkclass_detail_1
        
    comfirm_nightshift_1 = 0
    today_month_combobox_1.current(month)
    today_day_combobox_1.current(day)
    if kanto_mode == 0:
        i = 0
        u = 1
    elif kanto_mode == 1:
        i = 0
        u = 0
    time_h_num_combobox_1.current(i)
    time_m_num_combobox_1.current(u)
    
    comfirm_month_1 = int(today_month_combobox_1.get())
    comfirm_day_1 = int(today_day_combobox_1.get())
    mat_num_combobox_1.set("")
    wark_num_combobox_1.set("")
    warkclass_num_combobox_1.set("")
    endtime_h_num_combobox_1.set("")
    endtime_m_num_combobox_1.set("")
    total_label_v1.set("")
    break_listbox_1.set("")
    warkclass_detail_entry_1.delete(0,tk.END)
    checkbutton_var_1.set("0")
    
       
def row2clear():
    global comfirm_mat_num_2
    global comfirm_wark_num_2
    global comfirm_start_time_h_2
    global comfirm_start_time_m_2
    global comfirm_end_time_h_2
    global comfirm_end_time_m_2
    global comfirm_nightshift_2
    global comfirm_warkclass_detail_2
    global comfirm_month_2
    global comfirm_day_2    
    global today_month_combobox_2
    global today_day_combobox_2
    global comfirm_break_time_2
    global end_time_h2
    global end_time_m2
    global total_time_2

    if 'comfirm_mat_num_2' in globals():
        del comfirm_mat_num_2
    if 'comfirm_wark_num_2' in globals():
        del comfirm_wark_num_2
    if 'comfirm_start_time_h_2' in globals():
        del comfirm_start_time_h_2
    if 'comfirm_start_time_m_2' in globals():
        del comfirm_start_time_m_2
    if 'comfirm_end_time_h_2' in globals():
        del comfirm_end_time_h_2
    if 'comfirm_end_time_m_2' in globals():
        del comfirm_end_time_m_2
    if 'total_time_2' in globals():
        total_time_2 = 0
        

    if 'comfirm_break_time_2' in globals():
        comfirm_break_time_2 = 0
    if 'comfirm_warkclass_num_2' in globals():
        comfirm_warkclass_num_2 = 0
    if 'end_time_h2' in globals():
        del end_time_h2
    if 'end_time_m2' in globals():
        del end_time_m2
    if 'comfirm_warkclass_detail_2' in globals():
        del comfirm_warkclass_detail_2

    comfirm_nightshift_2 = 0
    today_month_combobox_2.current(month)
    today_day_combobox_2.current(day)
    time_h_num_combobox_2.current(0)
    time_m_num_combobox_2.current(0)
    comfirm_month_2 = int(today_month_combobox_2.get())
    comfirm_day_2 = int(today_day_combobox_2.get())
    mat_num_combobox_2.set("")
    wark_num_combobox_2.set("")
    warkclass_num_combobox_2.set("")
    time_h_num_combobox_2.set("")
    time_m_num_combobox_2.set("")
    endtime_h_num_combobox_2.set("")
    endtime_m_num_combobox_2.set("")
    break_listbox_2.set("")
    total_label_v2.set("")
    warkclass_detail_entry_2.delete(0,tk.END)
    checkbutton_var_2.set("0")

def row3clear():
    global comfirm_mat_num_3
    global comfirm_start_time_h_3
    global comfirm_wark_num_3
    global comfirm_start_time_m_3
    global comfirm_end_time_h_3
    global comfirm_end_time_m_3
    global comfirm_nightshift_3
    global comfirm_warkclass_detail_3
    global comfirm_month_3
    global comfirm_day_3    
    global today_month_combobox_3
    global today_day_combobox_3
    global comfirm_break_time_3
    global end_time_h3
    global end_time_m3
    global total_time_3
    if 'comfirm_mat_num_3' in globals():
        del comfirm_mat_num_3
    if 'comfirm_wark_num_3' in globals():
        del comfirm_wark_num_3
    if 'comfirm_start_time_h_3' in globals():
        del comfirm_start_time_h_3
    if 'comfirm_start_time_m_3' in globals():
        del comfirm_start_time_m_3
    if 'comfirm_end_time_h_3' in globals():
        del comfirm_end_time_h_3
    if 'comfirm_end_time_m_3' in globals():
        del comfirm_end_time_m_3
    if 'total_time_3' in globals():
        total_time_3 = 0
    if 'comfirm_break_time_3' in globals():
        comfirm_break_time_3 = 0
    if 'comfirm_warkclass_num_3' in globals():
        comfirm_warkclass_num_3 = 0
    if 'end_time_h3' in globals():
        del end_time_h3
    if 'end_time_m3' in globals():
        del end_time_m3
    if 'comfirm_warkclass_detail_3' in globals():
        del comfirm_warkclass_detail_3

    comfirm_nightshift_3 = 0
    today_month_combobox_3.current(month)
    today_day_combobox_3.current(day)
    time_h_num_combobox_3.current(0)
    time_m_num_combobox_3.current(0)
    comfirm_month_3 = int(today_month_combobox_3.get())
    comfirm_day_3 = int(today_day_combobox_3.get())
    mat_num_combobox_3.set("")
    wark_num_combobox_3.set("")
    warkclass_num_combobox_3.set("")
    time_h_num_combobox_3.set("")
    time_m_num_combobox_3.set("")
    endtime_h_num_combobox_3.set("")
    endtime_m_num_combobox_3.set("")
    break_listbox_3.set("")
    total_label_v3.set("")
    warkclass_detail_entry_3.delete(0,tk.END)
    checkbutton_var_3.set("0")

def row4clear():
    global comfirm_mat_num_4
    global comfirm_wark_num_4
    global comfirm_start_time_h_4
    global comfirm_start_time_m_4
    global comfirm_end_time_h_4
    global comfirm_end_time_m_4
    global comfirm_nightshift_4
    global comfirm_warkclass_detail_4
    global comfirm_month_4
    global comfirm_day_4   
    global today_month_combobox_4
    global today_day_combobox_4
    global comfirm_break_time_4
    global end_time_h4
    global end_time_m4
    global total_time_4
    if 'comfirm_mat_num_4' in globals():
        del comfirm_mat_num_4
    if 'comfirm_wark_num_4' in globals():
        del comfirm_wark_num_4
    if 'comfirm_start_time_h_4' in globals():
        del comfirm_start_time_h_4
    if 'comfirm_start_time_m_4' in globals():
        del comfirm_start_time_m_4
    if 'comfirm_end_time_h_4' in globals():
        del comfirm_end_time_h_4
    if 'comfirm_end_time_m_4' in globals():
        del comfirm_end_time_m_4
    if 'total_time_4' in globals():
        total_time_4 = 0
    if 'comfirm_break_time_4' in globals():
        comfirm_break_time_4 = 0
    if 'comfirm_warkclass_num_4' in globals():
        comfirm_warkclass_num_4 = 0
    if 'end_time_h4' in globals():
        del end_time_h4
    if 'end_time_m4' in globals():
        del end_time_m4
    if 'comfirm_warkclass_detail_4' in globals():
        del comfirm_warkclass_detail_4

    comfirm_nightshift_4 = 0
    today_month_combobox_4.current(month)
    today_day_combobox_4.current(day)
    time_h_num_combobox_4.current(0)
    time_m_num_combobox_4.current(0)
    comfirm_month_4 = int(today_month_combobox_4.get())
    comfirm_day_4 = int(today_day_combobox_4.get())
    wark_num_combobox_4.set("")
    warkclass_num_combobox_4.set("")
    mat_num_combobox_4.set("")
    time_h_num_combobox_4.set("")
    time_m_num_combobox_4.set("")
    endtime_h_num_combobox_4.set("")
    endtime_m_num_combobox_4.set("")
    break_listbox_4.set("")
    total_label_v4.set("")
    warkclass_detail_entry_4.delete(0,tk.END)
    checkbutton_var_4.set("0")

def row5clear():
    global comfirm_mat_num_5
    global comfirm_wark_num_5
    global comfirm_start_time_h_5
    global comfirm_start_time_m_5
    global comfirm_end_time_h_5
    global comfirm_end_time_m_5
    global comfirm_nightshift_5
    global comfirm_warkclass_detail_5
    global comfirm_month_5
    global comfirm_day_5    
    global today_month_combobox_5
    global today_day_combobox_5
    global comfirm_break_time_5
    global end_time_h5
    global end_time_m5
    global total_time_5
    if 'comfirm_mat_num_5' in globals():
        del comfirm_mat_num_5
    if 'comfirm_wark_num_5' in globals():
        del comfirm_wark_num_5
    if 'comfirm_start_time_h_5' in globals():
        del comfirm_start_time_h_5
    if 'comfirm_start_time_m_5' in globals():
        del comfirm_start_time_m_5
    if 'comfirm_end_time_h_5' in globals():
        del comfirm_end_time_h_5
    if 'comfirm_end_time_m_5' in globals():
        del comfirm_end_time_m_5
    if 'total_time_5' in globals():
        total_time_5 = 0
    if 'comfirm_break_time_5' in globals():
        comfirm_break_time_5 = 0
    if 'comfirm_warkclass_num_5' in globals():
        comfirm_warkclass_num_5 = 0
    if 'end_time_h5' in globals():
        del end_time_h5
    if 'end_time_m5' in globals():
        del end_time_m5
    if 'comfirm_warkclass_detail_5' in globals():
        del comfirm_warkclass_detail_5

    comfirm_nightshift_5 = 0
    today_month_combobox_5.current(month)
    today_day_combobox_5.current(day)
    time_h_num_combobox_5.current(0)
    time_m_num_combobox_5.current(0)
    comfirm_month_5 = int(today_month_combobox_5.get())
    comfirm_day_5 = int(today_day_combobox_5.get())
    mat_num_combobox_5.set("")
    wark_num_combobox_5.set("")
    warkclass_num_combobox_5.set("")    
    time_h_num_combobox_5.set("")
    time_m_num_combobox_5.set("")
    endtime_h_num_combobox_5.set("")
    endtime_m_num_combobox_5.set("")
    break_listbox_5.set("")    
    total_label_v5.set("")
    warkclass_detail_entry_5.delete(0,tk.END)
    checkbutton_var_5.set("0")

def row6clear():
    global comfirm_mat_num_6
    global comfirm_wark_num_6
    global comfirm_start_time_h_6
    global comfirm_start_time_m_6
    global comfirm_end_time_h_6
    global comfirm_end_time_m_6
    global comfirm_nightshift_6
    global comfirm_warkclass_detail_6
    global comfirm_month_6
    global comfirm_day_6    
    global today_month_combobox_6
    global today_day_combobox_6
    global comfirm_break_time_6
    global end_time_h6
    global end_time_m6
    global total_time_6
    if 'comfirm_mat_num_6' in globals():
        del comfirm_mat_num_6
    if 'comfirm_wark_num_6' in globals():
        del comfirm_wark_num_6
    if 'comfirm_start_time_h_6' in globals():
        del comfirm_start_time_h_6
    if 'comfirm_start_time_m_6' in globals():
        del comfirm_start_time_m_6
    if 'comfirm_end_time_h_6' in globals():
        del comfirm_end_time_h_6
    if 'comfirm_end_time_m_6' in globals():
        del comfirm_end_time_m_6
    if 'total_time_6' in globals():
        total_time_6 = 0
    if 'comfirm_break_time_6' in globals():
        comfirm_break_time_6 = 0
    if 'comfirm_warkclass_num_6' in globals():
        comfirm_warkclass_num_6 = 0
    if 'end_time_h6' in globals():
        del end_time_h6
    if 'end_time_m6' in globals():
        del end_time_m6
    if 'comfirm_warkclass_detail_6' in globals():
        del comfirm_warkclass_detail_6

    comfirm_nightshift_6 = 0
    today_month_combobox_6.current(month)
    today_day_combobox_6.current(day)
    time_h_num_combobox_6.current(0)
    time_m_num_combobox_6.current(0)
    comfirm_month_6 = int(today_month_combobox_6.get())
    comfirm_day_6 = int(today_day_combobox_6.get())
    wark_num_combobox_6.set("")
    warkclass_num_combobox_6.set("")    
    mat_num_combobox_6.set("")
    time_h_num_combobox_6.set("")
    time_m_num_combobox_6.set("")
    endtime_h_num_combobox_6.set("")
    endtime_m_num_combobox_6.set("")
    break_listbox_6.set("")    
    total_label_v6.set("")
    warkclass_detail_entry_6.delete(0,tk.END)
    checkbutton_var_6.set("0")

def row7clear():
    global comfirm_mat_num_7
    global comfirm_wark_num_7
    global comfirm_start_time_h_7
    global comfirm_start_time_m_7
    global comfirm_end_time_h_7
    global comfirm_end_time_m_7
    global comfirm_nightshift_7
    global comfirm_warkclass_detail_7
    global comfirm_month_7
    global comfirm_day_7    
    global today_month_combobox_7
    global today_day_combobox_7
    global comfirm_break_time_7
    global end_time_h7
    global end_time_m7
    global total_time_7
    if 'comfirm_mat_num_7' in globals():
        del comfirm_mat_num_7
    if 'comfirm_wark_num_7' in globals():
        del comfirm_wark_num_7
    if 'comfirm_start_time_h_7' in globals():
        del comfirm_start_time_h_7
    if 'comfirm_start_time_m_7' in globals():
        del comfirm_start_time_m_7
    if 'comfirm_end_time_h_7' in globals():
        del comfirm_end_time_h_7
    if 'comfirm_end_time_m_7' in globals():
        del comfirm_end_time_m_7
    if 'total_time_7' in globals():
        total_time_7 = 0
    if 'comfirm_break_time_7' in globals():
        comfirm_break_time_7 = 0
    if 'comfirm_warkclass_num_7' in globals():
        comfirm_warkclass_num_7 = 0
    if 'end_time_h7' in globals():
        del end_time_h7
    if 'end_time_m7' in globals():
        del end_time_m7
    if 'comfirm_warkclass_detail_7' in globals():
        del comfirm_warkclass_detail_7

    comfirm_nightshift_7 = 0
    today_month_combobox_7.current(month)
    today_day_combobox_7.current(day)
    time_h_num_combobox_7.current(0)
    time_m_num_combobox_7.current(0)
    comfirm_month_7 = int(today_month_combobox_7.get())
    comfirm_day_7 = int(today_day_combobox_7.get())
    mat_num_combobox_7.set("")
    wark_num_combobox_7.set("")
    warkclass_num_combobox_7.set("")    
    time_h_num_combobox_7.set("")
    time_m_num_combobox_7.set("")
    endtime_h_num_combobox_7.set("")
    endtime_m_num_combobox_7.set("")
    break_listbox_7.set("")    
    total_label_v7.set("")
    warkclass_detail_entry_7.delete(0,tk.END)
    checkbutton_var_7.set("0")

def row8clear():
    global comfirm_mat_num_8
    global comfirm_wark_num_8
    global comfirm_start_time_h_8
    global comfirm_start_time_m_8
    global comfirm_end_time_h_8
    global comfirm_end_time_m_8
    global comfirm_nightshift_8
    global comfirm_warkclass_detail_8
    global comfirm_month_8
    global comfirm_day_8    
    global today_month_combobox_8
    global today_day_combobox_8
    global comfirm_break_time_8
    global end_time_h8
    global end_time_m8
    global total_time_8
    if 'comfirm_mat_num_8' in globals():
        del comfirm_mat_num_8
    if 'comfirm_wark_num_8' in globals():
        del comfirm_wark_num_8
    if 'comfirm_start_time_h_8' in globals():
        del comfirm_start_time_h_8
    if 'comfirm_start_time_m_8' in globals():
        del comfirm_start_time_m_8
    if 'comfirm_end_time_h_8' in globals():
        del comfirm_end_time_h_8
    if 'comfirm_end_time_m_8' in globals():
        del comfirm_end_time_m_8
    if 'total_time_8' in globals():
        total_time_8 = 0
    if 'comfirm_break_time_8' in globals():
        comfirm_break_time_8 = 0
    if 'comfirm_warkclass_num_8' in globals():
        comfirm_warkclass_num_8 = 0
    if 'end_time_h8' in globals():
        del end_time_h8
    if 'end_time_m8' in globals():
        del end_time_m8
    if 'comfirm_warkclass_detail_8' in globals():
        del comfirm_warkclass_detail_8

    comfirm_nightshift_8 = 0
    today_month_combobox_8.current(month)
    today_day_combobox_8.current(day)
    time_h_num_combobox_8.current(0)
    time_m_num_combobox_8.current(0)
    comfirm_month_8 = int(today_month_combobox_8.get())
    comfirm_day_8 = int(today_day_combobox_8.get())
    wark_num_combobox_8.set("")
    warkclass_num_combobox_8.set("")
    mat_num_combobox_8.set("")
    time_h_num_combobox_8.set("")
    time_m_num_combobox_8.set("")
    endtime_h_num_combobox_8.set("")
    endtime_m_num_combobox_8.set("")
    break_listbox_8.set("")
    total_label_v8.set("")
    warkclass_detail_entry_8.delete(0,tk.END)
    checkbutton_var_8.set("0")

def excel_output(today, cmn, cwcn, cwn, cst_h, cst_m, cet_h, cet_m, total):
    start_time = datetime.time(cst_h,cst_m)
    end_time = datetime.time(cet_h,cet_m)
    comfirm = [today, comfirm_name, cmn, cwcn, cwn, start_time, end_time, total, overwark, nightshift_x]
    return comfirm

def excel_output_elif(today, cwcn, cwn, cst_h, cst_m, cet_h, cet_m, total):
    start_time = datetime.time(cst_h,cst_m)
    end_time = datetime.time(cet_h,cet_m)
    comfirm = [today, comfirm_name,  "", cwcn, cwn, start_time, end_time, total, overwark, nightshift_x]
    return comfirm

def set1():
    global Call_Empty
    
    Call_Empty = 0

    if 'comfirm_name' in globals():
        if'comfirm_start_time_h_1' in globals() and 'comfirm_start_time_m_1' in globals() and 'comfirm_end_time_h_1' in globals() and 'comfirm_end_time_m_1' in globals():
            wb2 = openpyxl.load_workbook(failpass2, keep_vba=True)
            ws2 = wb2.worksheets[0]
            wb3 = openpyxl.load_workbook(failpass3, keep_vba=True)
            ws3 = wb3.worksheets[0]
            if total_time >= 480:
                if overwark_check_mode == 1:
                    i = cell_check(ws3['A'], ws3)
                    list_i = [today, comfirm_name]
                    b = 1
                    for row_i in list_i:
                        ws3.cell(row=i,column=b).value = row_i
                        b+=1
                    try:
                        wb3.save(failpass3)
                    except:
                        messagebox.showwarning("確認","登録ファイルが開かれています。")
                        wb3.save(failpass3)

                Call_Empty = cell_check(ws2['E'], ws2)
                i = 0
                for cell_i in ws2['E']:
                    i += 1
                    if Call_Empty == i:
                        if cell_i.value == None:
                            break
                        else:
                            Call_Empty +=1
                            break

                check_column_1()            

            else:
                MessageBox = messagebox.askyesno("確認","合計時間が480分を満たしていません\n登録してもよろしいでしょうか？")

                if MessageBox == True:
                    if overwark_check_mode == 1:
                        i = cell_check(ws3['A'], ws3)
                        list_i = [today, comfirm_name,]
                        b = 1
                        for row_i in list_i:
                            ws3.cell(row=i,column=b).value = row_i
                            b+=1
                        wb3.save(failpass3)
                    Call_Empty = cell_check(ws2['E'], ws2)
                    check_column_1()
                
        else:
            messagebox.showwarning("確認","時間を選択してください")
    else:
        messagebox.showwarning("確認","名前を選択してください")
    
def check_column_1():
    global fix_column
    if'comfirm_start_time_h_1' in globals() and 'comfirm_start_time_m_1' in globals() and 'comfirm_end_time_h_1' in globals() and 'comfirm_end_time_m_1' in globals():    
        if fix_mode == 1:
            fix_column = 1
        if'comfirm_warkclass_num_1' in globals() and 'comfirm_wark_num_1' in globals():
            check_column_2()
        else:
            messagebox.showwarning("確認","1行目 入力されていない箇所があります。")
       
def check_column_2():
    global fix_column
    if'comfirm_start_time_h_2' in globals() and 'comfirm_start_time_m_2' in globals() and 'comfirm_end_time_h_2' in globals() and 'comfirm_end_time_m_2' in globals():
        if fix_mode == 1:
            fix_column = 2
        if 'comfirm_warkclass_num_2' in globals() and 'comfirm_wark_num_2' in globals():
            check_column_3()
        else:
            messagebox.showwarning("確認","2行目 入力されていない箇所があります。")
    else:
        row1comfirm()
       
def check_column_3():
    global fix_column

    if'comfirm_start_time_h_3' in globals() and 'comfirm_start_time_m_3' in globals() and 'comfirm_end_time_h_3' in globals() and 'comfirm_end_time_m_3' in globals():
        if fix_mode == 1:
            fix_column = 3
        if 'comfirm_warkclass_num_3' in globals() and 'comfirm_wark_num_3' in globals():
            check_column_4()
        else:
            messagebox.showwarning("確認","3行目 入力されていない箇所があります。")
    else:
        row1comfirm()
       
def check_column_4():
    global fix_column
    if'comfirm_start_time_h_4' in globals() and 'comfirm_start_time_m_4' in globals() and 'comfirm_end_time_h_4' in globals() and 'comfirm_end_time_m_4' in globals():
        if fix_mode == 1:
            fix_column = 4
        if 'comfirm_warkclass_num_4' in globals() and 'comfirm_wark_num_4' in globals():
            check_column_5()
        else:
            messagebox.showwarning("確認","4行目 入力されていない箇所があります。")
    else:
        row1comfirm()
           
def check_column_5():
    global fix_column
    if'comfirm_start_time_h_5' in globals() and 'comfirm_start_time_m_5' in globals() and 'comfirm_end_time_h_5' in globals() and 'comfirm_end_time_m_5' in globals():
        if fix_mode == 1:
            fix_column = 5
        if 'comfirm_warkclass_num_5' in globals() and 'comfirm_wark_num_5' in globals():
            check_column_6()
        else:
            messagebox.showwarning("確認","5行目 入力されていない箇所があります。")
    else:
        row1comfirm()
       
def check_column_6():
    global fix_column
    if'comfirm_start_time_h_6' in globals() and 'comfirm_start_time_m_6' in globals() and 'comfirm_end_time_h_6' in globals() and 'comfirm_end_time_m_6' in globals():
        if fix_mode == 1:
            fix_column = 6
        if 'comfirm_warkclass_num_6' in globals() and 'comfirm_wark_num_6' in globals():
            check_column_7()
        else:
            messagebox.showwarning("確認","6行目 入力されていない箇所があります。")
    else:
        row1comfirm()
       
def check_column_7():
    global fix_column
    if'comfirm_start_time_h_7' in globals() and 'comfirm_start_time_m_7' in globals() and 'comfirm_end_time_h_7' in globals() and 'comfirm_end_time_m_7' in globals():
        if fix_mode == 1:
            fix_column = 7
        if 'comfirm_warkclass_num_7' in globals() and 'comfirm_wark_num_7' in globals():
            check_column_8()
        else:
            messagebox.showwarning("確認","7行目 入力されていない箇所があります。")
    else:
        row1comfirm()
       
def check_column_8():
    global fix_column
    if'comfirm_start_time_h_8' in globals() and 'comfirm_start_time_m_8' in globals() and 'comfirm_end_time_h_8' in globals() and 'comfirm_end_time_m_8' in globals():
        if fix_mode == 1:

            fix_column = 8
        if 'comfirm_warkclass_num_8' in globals() and 'comfirm_wark_num_8' in globals():
            row1comfirm()
        else:
            messagebox.showwarning("確認","8行目 入力されていない箇所があります。")
    else:
        row1comfirm()                
       
def row1comfirm():
    global All_contents
    global overwark
    global nightshift_x
    global fix_column
    global fix_endrow

    All_contents_list = []

    # if fix_mode == 1:
    #     wb2 = openpyxl.load_workbook(failpass2, keep_vba=True)
    #     ws2 = wb2.worksheets[0]

    #     Call_Empty = fix_fastrow

    if 'comfirm_end_time_m_2' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""
    if 'comfirm_end_time_m_2' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""

    if 'comfirm_warkclass_detail_1' in globals() and 'comfirm_warkclass_num_1' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_1 + comfirm_warkclass_detail_1
    elif 'comfirm_warkclass_num_1' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_1 + ""

    if 'comfirm_start_time_h_1' in globals() and 'comfirm_start_time_m_1' in globals() and 'comfirm_end_time_h_1' in globals() and 'comfirm_end_time_m_1' in globals():
        if 'comfirm_mat_num_1' in globals() and 'comfirm_wark_num_1' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_1),
                                    int(comfirm_day_1)), 
                                    comfirm_mat_num_1, 
                                    comfirm_warkclass_detail, 
                                    comfirm_wark_num_1, 
                                    comfirm_start_time_h_1, 
                                    comfirm_start_time_m_1, 
                                    comfirm_end_time_h_1, 
                                    comfirm_end_time_m_1, 
                                    total_time_1
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_1),
                                    int(comfirm_day_1)), 
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_1, 
                                    comfirm_start_time_h_1, 
                                    comfirm_start_time_m_1, 
                                    comfirm_end_time_h_1, 
                                    comfirm_end_time_m_1, 
                                    total_time_1
                                    )
        All_contents.append(All_contents_list)
        row2comfirm()          

def row2comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x

    All_contents_list = []

    if 'comfirm_end_time_m_3' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""

    if 'comfirm_end_time_m_3' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_2' in globals() and 'comfirm_warkclass_num_2' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_2 + comfirm_warkclass_detail_2
    elif 'comfirm_warkclass_num_2' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_2 + ""      

    if 'comfirm_start_time_h_2' in globals() and 'comfirm_start_time_m_2' in globals() and 'comfirm_end_time_h_2' in globals() and 'comfirm_end_time_m_2' in globals():
        if 'comfirm_mat_num_2' in globals() and 'comfirm_wark_num_2' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_2),
                                    int(comfirm_day_2)), 
                                    comfirm_mat_num_2, 
                                    comfirm_warkclass_detail, 
                                    comfirm_wark_num_2,
                                    comfirm_start_time_h_2, 
                                    comfirm_start_time_m_2, 
                                    comfirm_end_time_h_2, 
                                    comfirm_end_time_m_2, 
                                    total_time_2
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_2),
                                    int(comfirm_day_2)), 
                                    comfirm_warkclass_detail, 
                                    comfirm_wark_num_2, 
                                    comfirm_start_time_h_2, 
                                    comfirm_start_time_m_2, 
                                    comfirm_end_time_h_2, 
                                    comfirm_end_time_m_2, 
                                    total_time_2
                                    )
        All_contents.append(All_contents_list)
        row3comfirm()
    else:
        save_contens(All_contents)

        
def row3comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x

    All_contents_list = []

    if 'comfirm_end_time_m_4' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""
    
    if 'comfirm_end_time_m_4' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_3' in globals() and 'comfirm_warkclass_num_3' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_3 + comfirm_warkclass_detail_3
    elif 'comfirm_warkclass_num_3' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_3 + ""

    if 'comfirm_start_time_h_3' in globals() and 'comfirm_start_time_m_3' in globals() and 'comfirm_end_time_h_3' in globals() and 'comfirm_end_time_m_3' in globals():
        if 'comfirm_mat_num_3' in globals() and 'comfirm_wark_num_3' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_3),
                                    int(comfirm_day_3)), 
                                    comfirm_mat_num_3, 
                                    comfirm_warkclass_detail, 
                                    comfirm_wark_num_3, 
                                    comfirm_start_time_h_3, 
                                    comfirm_start_time_m_3, 
                                    comfirm_end_time_h_3, 
                                    comfirm_end_time_m_3, 
                                    total_time_3
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_3),
                                    int(comfirm_day_3)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_3,
                                    comfirm_start_time_h_3,
                                    comfirm_start_time_m_3,
                                    comfirm_end_time_h_3,
                                    comfirm_end_time_m_3,
                                    total_time_3
                                    )
        All_contents.append(All_contents_list)
        row4comfirm()
    else:
        save_contens(All_contents)

def row4comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x

    All_contents_list = []

    if 'comfirm_end_time_m_5' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""

    if 'comfirm_end_time_m_5' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_4' in globals() and 'comfirm_warkclass_num_4' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_4 + comfirm_warkclass_detail_4
    elif 'comfirm_warkclass_num_4' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_4 + ""

    if 'comfirm_start_time_h_4' in globals() and 'comfirm_start_time_m_4' in globals() and 'comfirm_end_time_h_4' in globals() and 'comfirm_end_time_m_4' in globals():
        if 'comfirm_mat_num_4' in globals() and 'comfirm_wark_num_4' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_4),
                                    int(comfirm_day_4)),
                                    comfirm_mat_num_4,
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_4,
                                    comfirm_start_time_h_4,
                                    comfirm_start_time_m_4,
                                    comfirm_end_time_h_4,
                                    comfirm_end_time_m_4,
                                    total_time_4
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_4),
                                    int(comfirm_day_4)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_4,
                                    comfirm_start_time_h_4,
                                    comfirm_start_time_m_4,
                                    comfirm_end_time_h_4,
                                    comfirm_end_time_m_4,
                                    total_time_4
                                    )
        All_contents.append(All_contents_list)
        # Call_Empty+=1
        row5comfirm()    
    else:
        messagebox.showwarning("確認","登録しました。")
        allclear()
    
def row5comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x
    All_contents_list = []
    if 'comfirm_end_time_m_6' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""
    
    if 'comfirm_end_time_m_6' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_5' in globals() and 'comfirm_warkclass_num_5' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_5 + comfirm_warkclass_detail_5   
    elif 'comfirm_warkclass_num_5' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_5 + ""

    if 'comfirm_start_time_h_5' in globals() and 'comfirm_start_time_m_5' in globals() and 'comfirm_end_time_h_5' in globals() and 'comfirm_end_time_m_5' in globals():
        if 'comfirm_mat_num_5' in globals() and 'comfirm_wark_num_5' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_5),
                                    int(comfirm_day_5)),
                                    comfirm_mat_num_5,
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_5,
                                    comfirm_start_time_h_5,
                                    comfirm_start_time_m_5,
                                    comfirm_end_time_h_5,
                                    comfirm_end_time_m_5,
                                    total_time_5
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_5),
                                    int(comfirm_day_5)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_5,
                                    comfirm_start_time_h_5,
                                    comfirm_start_time_m_5,
                                    comfirm_end_time_h_5,
                                    comfirm_end_time_m_5,
                                    total_time_5
                                    )
        All_contents.append(All_contents_list)
        # Call_Empty+=1
        row6comfirm()
    else:
        save_contens(All_contents)
                
def row6comfirm():
    global All_contents
    global Call_Empty
    global overwark 
    global nightshift_x
    All_contents_list = []
    if 'comfirm_end_time_m_7' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""
    if 'comfirm_end_time_m_7' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_6' in globals() and 'comfirm_warkclass_num_6' in globals():
        comfirm_warkclass_detail  = comfirm_warkclass_num_6 + comfirm_warkclass_detail_6
    elif 'comfirm_warkclass_num_6' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_6 + ""

    if 'comfirm_start_time_h_6' in globals() and 'comfirm_start_time_m_6' in globals() and 'comfirm_end_time_h_6' in globals() and 'comfirm_end_time_m_6' in globals():
        if 'comfirm_mat_num_6' in globals() and 'comfirm_wark_num_6' in globals():
           All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_6),
                                    int(comfirm_day_6)),
                                    comfirm_mat_num_6,
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_6,
                                    comfirm_start_time_h_6,
                                    comfirm_start_time_m_6,
                                    comfirm_end_time_h_6,
                                    comfirm_end_time_m_6,
                                    total_time_6
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_6),
                                    int(comfirm_day_6)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_6,
                                    comfirm_start_time_h_6,
                                    comfirm_start_time_m_6,
                                    comfirm_end_time_h_6,
                                    comfirm_end_time_m_6,
                                    total_time_6
                                    )
        All_contents.append(All_contents_list)
        # Call_Empty+=1
        row7comfirm()       
    else:
        save_contens(All_contents)

def row7comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x
    All_contents_list = []
    if 'comfirm_end_time_m_8' in globals():
        overwark = ""
    else:
        if overwark_time > 0:
            overwark = overwark_time
        else:
            overwark = ""

    if 'comfirm_end_time_m_8' in globals():
        nightshift_x = ""
    else:
        if comfirm_nightshift > 0:
            nightshift_x = comfirm_nightshift
        else:
            nightshift_x = ""
    
    if 'comfirm_warkclass_detail_7' in globals() and 'comfirm_warkclass_num_7' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_7 + comfirm_warkclass_detail_7
    elif 'comfirm_warkclass_num_7' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_7 + ""

    if 'comfirm_start_time_h_7' in globals() and 'comfirm_start_time_m_7' in globals() and 'comfirm_end_time_h_7' in globals() and 'comfirm_end_time_m_7' in globals():
        if 'comfirm_mat_num_7' in globals() and 'comfirm_wark_num_7' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_7),
                                    int(comfirm_day_7)),
                                    comfirm_mat_num_7,
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_7,
                                    comfirm_start_time_h_7,
                                    comfirm_start_time_m_7,
                                    comfirm_end_time_h_7,
                                    comfirm_end_time_m_7,
                                    total_time_7
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_7),
                                    int(comfirm_day_7)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_7,
                                    comfirm_start_time_h_7,
                                    comfirm_start_time_m_7,
                                    comfirm_end_time_h_7,
                                    comfirm_end_time_m_7,
                                    total_time_7
                                    )
        
        All_contents.append(All_contents_list)
        # Call_Empty+=1
        row8comfirm()        
    else:
        save_contens(All_contents)
 
def row8comfirm():
    global All_contents
    global Call_Empty
    global overwark
    global nightshift_x
    All_contents_list = []
    overwark = overwark_time
    nightshift_x = comfirm_nightshift

    if 'comfirm_warkclass_detail_8' in globals() and 'comfirm_warkclass_num_8' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_8 + comfirm_warkclass_detail_8
    elif 'comfirm_warkclass_num_8' in globals():
        comfirm_warkclass_detail = comfirm_warkclass_num_8 + ""

    if 'comfirm_start_time_h_8' in globals() and 'comfirm_start_time_m_8' in globals() and 'comfirm_end_time_h_8' in globals() and 'comfirm_end_time_m_8' in globals():
        if 'comfirm_mat_num_8' in globals() and 'comfirm_wark_num_8' in globals():
            All_contents_list = excel_output(datetime.date(ny,
                                    int(comfirm_month_8),
                                    int(comfirm_day_8)),
                                    comfirm_mat_num_8,
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_8,
                                    comfirm_start_time_h_8,
                                    comfirm_start_time_m_8,
                                    comfirm_end_time_h_8,
                                    comfirm_end_time_m_8,
                                    total_time_8
                                    )
        else:
            All_contents_list = excel_output_elif(datetime.date(ny,
                                    int(comfirm_month_8),
                                    int(comfirm_day_8)),
                                    comfirm_warkclass_detail,
                                    comfirm_wark_num_8,
                                    comfirm_start_time_h_8,
                                    comfirm_start_time_m_8,
                                    comfirm_end_time_h_8,
                                    comfirm_end_time_m_8,
                                    total_time_8
                                    )
    All_contents.append(All_contents_list)
    save_contens(All_contents)

def func1(event):
    global comfirm_name
    i = name_number.index(name_num_combobox.get())
    comfirm_name = str(name_number[i])
    name_ward_str_h1 = str(name_number[i])
    ward_var_1.set(name_ward_str_h1)

def today_month_1(event):
    comfirm_month_1 = today_month_combobox_1.get()
    return comfirm_month_1

def today_month_2(event):
    comfirm_month_2 = today_month_combobox_2.get()
    return comfirm_month_2

def today_month_3(event):
    comfirm_month_3 = today_month_combobox_3.get()
    return comfirm_month_3

def today_month_4(event):
    comfirm_month_4 = today_month_combobox_4.get()
    return comfirm_month_4

def today_month_5(event):
    comfirm_month_5 = today_month_combobox_5.get()
    return comfirm_month_5

def today_month_6(event):
    comfirm_month_6 = today_month_combobox_6.get()
    return comfirm_month_6

def today_month_7(event):
    comfirm_month_7 = today_month_combobox_7.get()
    return comfirm_month_7

def today_month_8(event):
    comfirm_month_8 = today_month_combobox_8.get()
    return comfirm_month_8

def today_day_1(event):
    global comfirm_day_1
    comfirm_day_1 = int(today_day_combobox_1.get())
    
def today_day_2(event):
    global comfirm_day_2
    comfirm_day_2 = int(today_day_combobox_2.get())

def today_day_3(event):
    global comfirm_day_3
    comfirm_day_3 = int(today_day_combobox_3.get())

def today_day_4(event):
    global comfirm_day_4
    comfirm_day_4 = int(today_day_combobox_4.get())

def today_day_5(event):
    global comfirm_day_5
    comfirm_day_5 = int(today_day_combobox_5.get())

def today_day_6(event):
    global comfirm_day_6
    comfirm_day_6 = int(today_day_combobox_6.get())

def today_day_7(event):
    global comfirm_day_7
    comfirm_day_7 = int(today_day_combobox_7.get())

def today_day_8(event):
    global comfirm_day_8
    comfirm_day_8 = int(today_day_combobox_8.get())



def mat_num_1(event):
    global comfirm_mat_num_1
    comfirm_mat_num_1 = mat_num_combobox_1.get()

def mat_num_2(event):
    global comfirm_mat_num_2
    comfirm_mat_num_2 = mat_num_combobox_2.get()

def mat_num_3(event):
    global comfirm_mat_num_3
    comfirm_mat_num_3 = mat_num_combobox_3.get()

def mat_num_4(event):
    global comfirm_mat_num_4
    comfirm_mat_num_4 = mat_num_combobox_4.get()

def mat_num_5(event):
    global comfirm_mat_num_5
    comfirm_mat_num_5 = mat_num_combobox_5.get()

def mat_num_6(event):
    global comfirm_mat_num_6
    comfirm_mat_num_6 = mat_num_combobox_6.get()

def mat_num_7(event):
    global comfirm_mat_num_7
    comfirm_mat_num_7 = mat_num_combobox_7.get()

def mat_num_8(event):
    global comfirm_mat_num_8
    comfirm_mat_num_8 = mat_num_combobox_8.get()
    
def wark_num_1(event):
    global comfirm_wark_num_1
    comfirm_wark_num_1 = wark_num_combobox_1.get()

def wark_num_2(event):
    global comfirm_wark_num_2
    comfirm_wark_num_2 = wark_num_combobox_2.get()

def wark_num_3(event):
    global comfirm_wark_num_3
    comfirm_wark_num_3 = wark_num_combobox_3.get()

def wark_num_4(event):
    global comfirm_wark_num_4
    comfirm_wark_num_4 = wark_num_combobox_4.get()

def wark_num_5(event):
    global comfirm_wark_num_5
    comfirm_wark_num_5 = wark_num_combobox_5.get()

def wark_num_6(event):
    global comfirm_wark_num_6
    comfirm_wark_num_6 = wark_num_combobox_6.get()

def wark_num_7(event):
    global comfirm_wark_num_7
    comfirm_wark_num_7 = wark_num_combobox_7.get()

def wark_num_8(event):
    global comfirm_wark_num_8
    comfirm_wark_num_8 = wark_num_combobox_8.get()

def warkclass_num_1(event):
    global comfirm_warkclass_num_1
    comfirm_warkclass_num_1 = warkclass_num_combobox_1.get()

def warkclass_num_2(event):
    global comfirm_warkclass_num_2
    comfirm_warkclass_num_2 = warkclass_num_combobox_2.get()


def warkclass_num_3(event):
    global comfirm_warkclass_num_3
    comfirm_warkclass_num_3 = warkclass_num_combobox_3.get()

def warkclass_num_4(event):
    global comfirm_warkclass_num_4
    comfirm_warkclass_num_4 = warkclass_num_combobox_4.get()

def warkclass_num_5(event):
    global comfirm_warkclass_num_5
    comfirm_warkclass_num_5 = warkclass_num_combobox_5.get()

def warkclass_num_6(event):
    global comfirm_warkclass_num_6
    comfirm_warkclass_num_6 = warkclass_num_combobox_6.get()

def warkclass_num_7(event):
    global comfirm_warkclass_num_7
    comfirm_warkclass_num_7 = warkclass_num_combobox_7.get()

def warkclass_num_8(event):
    global comfirm_warkclass_num_8
    comfirm_warkclass_num_8 = warkclass_num_combobox_8.get()

def warkclass_detail_get_1(event):
    global comfirm_warkclass_detail_1
    comfirm_warkclass_detail_1 = str(warkclass_detail_entry_1.get())
def warkclass_detail_get_1(event):
    global comfirm_warkclass_detail_1
    comfirm_warkclass_detail_1 = str(warkclass_detail_entry_1.get())

def warkclass_detail_get_2(event):
    global comfirm_warkclass_detail_2
    comfirm_warkclass_detail_2 = str(warkclass_detail_entry_2.get())

def warkclass_detail_get_3(event):
    global comfirm_warkclass_detail_3
    comfirm_warkclass_detail_3 = str(warkclass_detail_entry_3.get())

def warkclass_detail_get_4(event):
    global comfirm_warkclass_detail_4
    comfirm_warkclass_detail_4 = str(warkclass_detail_entry_4.get())

def warkclass_detail_get_5(event):
    global comfirm_warkclass_detail_5
    comfirm_warkclass_detail_5 = str(warkclass_detail_entry_5.get())

def warkclass_detail_get_6(event):
    global comfirm_warkclass_detail_6
    comfirm_warkclass_detail_6 = str(warkclass_detail_entry_6.get())

def warkclass_detail_get_7(event):
    global comfirm_warkclass_detail_7
    comfirm_warkclass_detail_7 = str(warkclass_detail_entry_7.get())

def warkclass_detail_get_8(event):
    global comfirm_warkclass_detail_8
    comfirm_warkclass_detail_8 = str(warkclass_detail_entry_8.get())

def row_copybutton_func_1():
    global comfirm_mat_num_2
    comfirm_mat_num_2 = comfirm_mat_num_1
    f = mat_number.index(str(comfirm_mat_num_1))
    mat_num_combobox_2.current(f)

def row_copybutton_func_2():
    global comfirm_mat_num_3
    comfirm_mat_num_3 = comfirm_mat_num_2
    f = mat_number.index(str(comfirm_mat_num_2))
    mat_num_combobox_3.current(f)

def row_copybutton_func_3():
    global comfirm_mat_num_4
    comfirm_mat_num_4 = comfirm_mat_num_3
    f = mat_number.index(str(comfirm_mat_num_3))
    mat_num_combobox_4.current(f)

def row_copybutton_func_4():
    global comfirm_mat_num_5
    comfirm_mat_num_5 = comfirm_mat_num_4
    f = mat_number.index(str(comfirm_mat_num_4))
    mat_num_combobox_5.current(f)

def row_copybutton_func_5():
    global comfirm_mat_num_6
    comfirm_mat_num_6 = comfirm_mat_num_5
    f = mat_number.index(str(comfirm_mat_num_5))
    mat_num_combobox_6.current(f)

def row_copybutton_func_6():
    global comfirm_mat_num_7
    comfirm_mat_num_7 = comfirm_mat_num_6
    f = mat_number.index(str(comfirm_mat_num_6))
    mat_num_combobox_7.current(f)

def row_copybutton_func_7():
    global comfirm_mat_num_8
    comfirm_mat_num_8 = comfirm_mat_num_7
    f = mat_number.index(str(comfirm_mat_num_7))
    mat_num_combobox_8.current(f)

def start_number_change_h1(event):
    global start_time_h1
    global comfirm_start_time_h_1
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h1 = comfirm_start_time_h_1 = int(time_h_num_combobox_1.get())

    if 'start_time_h1' in globals() and 'start_time_m1' in globals() and 'end_time_h1' in globals() and 'end_time_m1' in globals():
        total_time_1,plus = break_time_func(start_time_h1, start_time_m1, end_time_h1, end_time_m1,int(comfirm_month_1),int(comfirm_day_1),int(comfirm_break_time_1))
        comfirm_nightshift_1 = nightshift(plus)
        total_label_v1.set(str(total_time_1) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       

def start_number_change_h2(event):
    global start_time_h2
    global comfirm_start_time_h_2
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    start_time_h2 = comfirm_start_time_h_2 = int(time_h_num_combobox_2.get())

    if 'start_time_h2' in globals() and 'start_time_m2' in globals() and 'end_time_h2' in globals() and 'end_time_m2' in globals():
        total_time_2,plus = break_time_func(start_time_h2, start_time_m2, end_time_h2, end_time_m2,int(comfirm_month_2),int(comfirm_day_2),int(comfirm_break_time_2))
        comfirm_nightshift_2 = nightshift(plus)

        total_label_v2.set(str(total_time_2) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def start_number_change_h3(event):
    global start_time_h3
    global comfirm_start_time_h_3
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h3 = comfirm_start_time_h_3 = int(time_h_num_combobox_3.get())

    if 'start_time_h3' in globals() and 'start_time_m3' in globals() and 'end_time_h3' in globals() and 'end_time_m3' in globals():
        total_time_3,plus = break_time_func(start_time_h3, start_time_m3, end_time_h3, end_time_m3,int(comfirm_month_3),int(comfirm_day_3),int(comfirm_break_time_3))
        comfirm_nightshift_3 = nightshift(plus)

        total_label_v3.set(str(total_time_3) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def start_number_change_h4(event):
    global start_time_h4
    global comfirm_start_time_h_4
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h4 = comfirm_start_time_h_4 = int(time_h_num_combobox_4.get())

    if 'start_time_h4' in globals() and 'start_time_m4' in globals() and 'end_time_h4' in globals() and 'end_time_m4' in globals():
        total_time_4,plus = break_time_func(start_time_h4, start_time_m4, end_time_h4, end_time_m4,int(comfirm_month_4),int(comfirm_day_4),int(comfirm_break_time_4))    
        comfirm_nightshift_4 = nightshift(plus)

        total_label_v4.set(str(total_time_4) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def start_number_change_h5(event):
    global start_time_h5
    global comfirm_start_time_h_5
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h5 = comfirm_start_time_h_5 = int(time_h_num_combobox_5.get())
    
    if 'start_time_h5' in globals() and 'start_time_m5' in globals() and 'end_time_h5' in globals() and 'end_time_m5' in globals():
        total_time_5,plus = break_time_func(start_time_h5, start_time_m5, end_time_h5, end_time_m5,int(comfirm_month_5),int(comfirm_day_5),int(comfirm_break_time_5))    
        comfirm_nightshift_5 = nightshift(plus)

        total_label_v5.set(str(total_time_5) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()        

def start_number_change_h6(event):
    global start_time_h6
    global comfirm_start_time_h_6
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h6 = comfirm_start_time_h_6 = int(time_h_num_combobox_6.get())

    if 'start_time_h6' in globals() and 'start_time_m6' in globals() and 'end_time_h6' in globals() and 'end_time_m6' in globals():
        total_time_6,plus = break_time_func(start_time_h6, start_time_m6, end_time_h6, end_time_m6,int(comfirm_month_6),int(comfirm_day_6),int(comfirm_break_time_6))    

        comfirm_nightshift_6 = nightshift(plus)

        total_label_v6.set(str(total_time_6) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       

def start_number_change_h7(event):
    global start_time_h7
    global comfirm_start_time_h_7
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h7 = comfirm_start_time_h_7 = int(time_h_num_combobox_7.get())

    if 'start_time_h7' in globals() and 'start_time_m7' in globals() and 'end_time_h7' in globals() and 'end_time_m7' in globals():
        total_time_7,plus = break_time_func(start_time_h7, start_time_m7, end_time_h7, end_time_m7,int(comfirm_month_7),int(comfirm_day_7),int(comfirm_break_time_7))    

        comfirm_nightshift_7 = nightshift(plus)

        total_label_v7.set(str(total_time_7) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def start_number_change_h8(event):
    global start_time_h8
    global comfirm_start_time_h_8
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_h8 = comfirm_start_time_h_8 = int(time_h_num_combobox_8.get())

    if 'start_time_h8' in globals() and 'start_time_m8' in globals() and 'end_time_h8' in globals() and 'end_time_m8' in globals():
        total_time_8,plus = break_time_func(start_time_h8, start_time_m8, end_time_h8, end_time_m8,int(comfirm_month_8),int(comfirm_day_8),int(comfirm_break_time_8))
        
        comfirm_nightshift_8 = nightshift(plus)

        total_label_v8.set(str(total_time_8) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
                
def start_number_change_m1(event):
    global start_time_m1
    global comfirm_start_time_m_1
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m1 = comfirm_start_time_m_1 = int(time_m_num_combobox_1.get())
    
    if 'start_time_h1' in globals() and 'start_time_m1' in globals() and 'end_time_h1' in globals() and 'end_time_m1' in globals():
        total_time_1,plus = break_time_func(start_time_h1, start_time_m1, end_time_h1, end_time_m1,int(comfirm_month_1),int(comfirm_day_1),int(comfirm_break_time_1))
        comfirm_nightshift_1 = nightshift(plus)

        total_label_v1.set(str(total_time_1) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       

def start_number_change_m2(event):
    global start_time_m2
    global comfirm_start_time_m_2
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m2 = comfirm_start_time_m_2 = int(time_m_num_combobox_2.get())
    
    if 'start_time_h2' in globals() and 'start_time_m2' in globals() and 'end_time_h2' in globals() and 'end_time_m2' in globals():
        total_time_2,plus = break_time_func(start_time_h2, start_time_m2, end_time_h2, end_time_m2,int(comfirm_month_2),int(comfirm_day_2),int(comfirm_break_time_2))
        comfirm_nightshift_2 = nightshift(plus)

        total_label_v2.set(str(total_time_2) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       
        
def start_number_change_m3(event):
    global start_time_m3
    global comfirm_start_time_m_3
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m3 = comfirm_start_time_m_3 = int(time_m_num_combobox_3.get())

    if 'start_time_h3' in globals() and 'start_time_m3' in globals() and 'end_time_h3' in globals() and 'end_time_m3' in globals():
        total_time_3,plus = break_time_func(start_time_h3, start_time_m3, end_time_h3, end_time_m3,int(comfirm_month_3),int(comfirm_day_3),int(comfirm_break_time_3))
        comfirm_nightshift_3 = nightshift(plus)

        total_label_v3.set(str(total_time_3) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       
        
def start_number_change_m4(event):
    global start_time_m4
    global comfirm_start_time_m_4
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m4 = comfirm_start_time_m_4 = int(time_m_num_combobox_4.get())

    if 'start_time_h4' in globals() and 'start_time_m4' in globals() and 'end_time_h4' in globals() and 'end_time_m4' in globals():
        total_time_4,plus = break_time_func(start_time_h4, start_time_m4, end_time_h4, end_time_m4,int(comfirm_month_4),int(comfirm_day_4),int(comfirm_break_time_4))    
        comfirm_nightshift_4 = nightshift(plus)

        total_label_v4.set(str(total_time_4) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       
        
def start_number_change_m5(event):
    global start_time_m5
    global comfirm_start_time_m_5
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m5 = comfirm_start_time_m_5 = int(time_m_num_combobox_5.get())

    if 'start_time_h5' in globals() and 'start_time_m5' in globals() and 'end_time_h5' in globals() and 'end_time_m5' in globals():
        total_time_5,plus = break_time_func(start_time_h5, start_time_m5, end_time_h5, end_time_m5,int(comfirm_month_5),int(comfirm_day_5),int(comfirm_break_time_5))    
        comfirm_nightshift_5 = nightshift(plus)

        total_label_v5.set(str(total_time_5) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       
        
def start_number_change_m6(event):
    global start_time_m6
    global comfirm_start_time_m_6
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m6 = comfirm_start_time_m_6 = int(time_m_num_combobox_6.get())

    if 'start_time_h6' in globals() and 'start_time_m6' in globals() and 'end_time_h6' in globals() and 'end_time_m6' in globals():
        total_time_6,plus = break_time_func(start_time_h6, start_time_m6, end_time_h6, end_time_m6,int(comfirm_month_6),int(comfirm_day_6),int(comfirm_break_time_6))    

        comfirm_nightshift_6 = nightshift(plus)

        total_label_v6.set(str(total_time_6) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       

def start_number_change_m7(event):
    global start_time_m7
    global comfirm_start_time_m_7
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m7 = comfirm_start_time_m_7 = int(time_m_num_combobox_7.get())

    if 'start_time_h7' in globals() and 'start_time_m7' in globals() and 'end_time_h7' in globals() and 'end_time_m7' in globals():
        total_time_7,plus = break_time_func(start_time_h7, start_time_m7, end_time_h7, end_time_m7,int(comfirm_month_7),int(comfirm_day_7),int(comfirm_break_time_7))    

        comfirm_nightshift_7 = nightshift(plus)

        total_label_v7.set(str(total_time_7) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    # Calculation()       
        
def start_number_change_m8(event):
    global start_time_m8
    global comfirm_start_time_m_8
    global comfirm_nightshift
    global overwark_time
    global total_time

    start_time_m8 = comfirm_start_time_m_8 = int(time_m_num_combobox_8.get())

    if 'start_time_h8' in globals() and 'start_time_m8' in globals() and 'end_time_h8' in globals() and 'end_time_m8' in globals():
        total_time_8,plus = break_time_func(start_time_h8, start_time_m8, end_time_h8, end_time_m8,int(comfirm_month_8),int(comfirm_day_8),int(comfirm_break_time_8))
        
        comfirm_nightshift_8 = nightshift(plus)

        total_label_v8.set(str(total_time_8) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
                
def end_number_change_h1(event):
    global end_time_h1
    global start_time_h2
    global end_time_h2
    global comfirm_end_time_h_1
    global comfirm_start_time_h_2
    global comfirm_end_time_h_2
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v2.set(endtime_h_num_combobox_1.get())
    endtime_h_num_v2.set(endtime_h_num_combobox_1.get())
    
    end_time_h1 = end_time_h2 = start_time_h2 = comfirm_end_time_h_1 = comfirm_start_time_h_2 = comfirm_end_time_h_2 = int(endtime_h_num_combobox_1.get())
    
    if 'start_time_h1' in globals() and 'start_time_m1' in globals() and 'end_time_h1' in globals() and 'end_time_m1' in globals():
        total_time_1,plus = break_time_func(start_time_h1, start_time_m1, end_time_h1, end_time_m1,int(comfirm_month_1),int(comfirm_day_1),int(comfirm_break_time_1))
        comfirm_nightshift_1 = nightshift(plus)

        total_label_v1.set(str(total_time_1) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def end_number_change_h2(event):
    global end_time_h2
    global start_time_h3
    global end_time_h3
    global comfirm_end_time_h_2
    global comfirm_start_time_h_3
    global comfirm_end_time_h_3
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v3.set(endtime_h_num_combobox_2.get())
    endtime_h_num_v3.set(endtime_h_num_combobox_2.get())

    end_time_h2 = end_time_h3 = start_time_h3 = comfirm_end_time_h_2 = comfirm_start_time_h_3 = comfirm_end_time_h_3 = int(endtime_h_num_combobox_2.get())

    if 'start_time_h2' in globals() and 'start_time_m2' in globals() and 'end_time_h2' in globals() and 'end_time_m2' in globals():
        total_time_2,plus = break_time_func(start_time_h2, start_time_m2, end_time_h2, end_time_m2,int(comfirm_month_2),int(comfirm_day_2),int(comfirm_break_time_2))
        comfirm_nightshift_2 = nightshift(plus)

        total_label_v2.set(str(total_time_2) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def end_number_change_h3(event):
    global end_time_h3
    global start_time_h4
    global end_time_h4
    global comfirm_end_time_h_3
    global comfirm_start_time_h_4
    global comfirm_end_time_h_4
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v4.set(endtime_h_num_combobox_3.get())
    endtime_h_num_v4.set(endtime_h_num_combobox_3.get())

    end_time_h3 = end_time_h4 = start_time_h4 = comfirm_end_time_h_3 = comfirm_start_time_h_4 = comfirm_end_time_h_4 = int(endtime_h_num_combobox_3.get())
    
    if 'start_time_h3' in globals() and 'start_time_m3' in globals() and 'end_time_h3' in globals() and 'end_time_m3' in globals():
        total_time_3,plus = break_time_func(start_time_h3, start_time_m3, end_time_h3, end_time_m3,int(comfirm_month_3),int(comfirm_day_3),int(comfirm_break_time_3))
        comfirm_nightshift_3 = nightshift(plus)

        total_label_v3.set(str(total_time_3) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def end_number_change_h4(event):
    global end_time_h4
    global start_time_h5
    global end_time_h5
    global comfirm_end_time_h_4
    global comfirm_start_time_h_5
    global comfirm_end_time_h_5
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v5.set(endtime_h_num_combobox_4.get())
    endtime_h_num_v5.set(endtime_h_num_combobox_4.get())

    end_time_h4 = start_time_h5 = end_time_h5 = comfirm_end_time_h_4 = comfirm_start_time_h_5 = comfirm_end_time_h_5 = int(endtime_h_num_combobox_4.get())

    if 'start_time_h4' in globals() and 'start_time_m4' in globals() and 'end_time_h4' in globals() and 'end_time_m4' in globals():
        total_time_4,plus = break_time_func(start_time_h4, start_time_m4, end_time_h4, end_time_m4,int(comfirm_month_4),int(comfirm_day_4),int(comfirm_break_time_4))    
        comfirm_nightshift_4 = nightshift(plus)

        total_label_v4.set(str(total_time_4) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def end_number_change_h5(event):
    global end_time_h5
    global start_time_h6
    global end_time_h6
    global comfirm_end_time_h_5
    global comfirm_start_time_h_6
    global comfirm_end_time_h_6
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v6.set(endtime_h_num_combobox_5.get())
    endtime_h_num_v6.set(endtime_h_num_combobox_5.get())

    end_time_h5 = start_time_h6 = end_time_h6 = comfirm_end_time_h_5 = comfirm_start_time_h_6 = comfirm_end_time_h_6 = int(endtime_h_num_combobox_5.get())

    if 'start_time_h5' in globals() and 'start_time_m5' in globals() and 'end_time_h5' in globals() and 'end_time_m5' in globals():
        total_time_5,plus = break_time_func(start_time_h5, start_time_m5, end_time_h5, end_time_m5,int(comfirm_month_5),int(comfirm_day_5),int(comfirm_break_time_5))    
        comfirm_nightshift_5 = nightshift(plus)

        total_label_v5.set(str(total_time_5) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
        
def end_number_change_h6(event):
    global end_time_h6
    global start_time_h7
    global end_time_h7
    global comfirm_end_time_h_6
    global comfirm_start_time_h_7
    global comfirm_end_time_h_7
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v7.set(endtime_h_num_combobox_6.get())
    endtime_h_num_v7.set(endtime_h_num_combobox_6.get())

    end_time_h6 = start_time_h7 = end_time_h7 = comfirm_end_time_h_6 = comfirm_start_time_h_7 = comfirm_end_time_h_7 = int(endtime_h_num_combobox_6.get())

    if 'start_time_h6' in globals() and 'start_time_m6' in globals() and 'end_time_h6' in globals() and 'end_time_m6' in globals():
        total_time_6,plus = break_time_func(start_time_h6, start_time_m6, end_time_h6, end_time_m6,int(comfirm_month_6),int(comfirm_day_6),int(comfirm_break_time_6))    

        comfirm_nightshift_6 = nightshift(plus)

        total_label_v6.set(str(total_time_6) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def end_number_change_h7(event):
    global end_time_h7
    global start_time_h8
    global end_time_h8
    global comfirm_end_time_h_7
    global comfirm_start_time_h_8
    global comfirm_end_time_h_8
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_v8.set(endtime_h_num_combobox_7.get())
    endtime_h_num_v8.set(endtime_h_num_combobox_7.get())

    end_time_h7 = start_time_h8 = end_time_h8 = comfirm_end_time_h_7 = comfirm_start_time_h_8 = comfirm_end_time_h_8 = int(endtime_h_num_combobox_7.get())

    if 'start_time_h7' in globals() and 'start_time_m7' in globals() and 'end_time_h7' in globals() and 'end_time_m7' in globals():
        total_time_7,plus = break_time_func(start_time_h7, start_time_m7, end_time_h7, end_time_m7,int(comfirm_month_7),int(comfirm_day_7),int(comfirm_break_time_7))    

        comfirm_nightshift_7 = nightshift(plus)

        total_label_v7.set(str(total_time_7) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
          
def end_number_change_h8(event):
    global end_time_h8
    global comfirm_end_time_h_8
    global comfirm_nightshift
    global overwark_time
    global total_time
    
    time_h_num_str_h8 = endtime_h_num_combobox_8.get()
    end_time_h8 = comfirm_end_time_h_8 = int(endtime_h_num_combobox_8.get())

    if 'start_time_h8' in globals() and 'start_time_m8' in globals() and 'end_time_h8' in globals() and 'end_time_m8' in globals():
        total_time_8,plus = break_time_func(start_time_h8, start_time_m8, end_time_h8, end_time_m8,int(comfirm_month_8),int(comfirm_day_8),int(comfirm_break_time_8))
        
        comfirm_nightshift_8 = nightshift(plus)

        total_label_v8.set(str(total_time_8) + " 分")
        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation() 

def end_number_change_m1(event):
    global end_time_m1
    global start_time_m2
    global total_time_1
    global comfirm_end_time_m_1
    global comfirm_start_time_m_2
    global comfirm_nightshift_1
    global overwark_time
    global total_time
    global comfirm_nightshift
   
    time_m_num_v2.set(endtime_m_num_combobox_1.get())
    end_time_m1 = start_time_m2 = comfirm_end_time_m_1 = comfirm_start_time_m_2 = int(endtime_m_num_combobox_1.get())

    if 'start_time_h1' in globals() and 'start_time_m1' in globals() and 'end_time_h1' in globals() and 'end_time_m1' in globals():
        total_time_1,plus = break_time_func(start_time_h1, start_time_m1, end_time_h1, end_time_m1,int(comfirm_month_1),int(comfirm_day_1),int(comfirm_break_time_1))
        comfirm_nightshift_1 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v1.set(str(total_time_1) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
     
def end_number_change_m2(event):
    global end_time_m2
    global start_time_m3
    global total_time_2
    global comfirm_end_time_m_2
    global comfirm_start_time_m_3
    global comfirm_nightshift_2
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v3.set(endtime_m_num_combobox_2.get())
    end_time_m2 = start_time_m3 = comfirm_end_time_m_2 = comfirm_start_time_m_3 = int(endtime_m_num_combobox_2.get())

    if 'start_time_h2' in globals() and 'start_time_m2' in globals() and 'end_time_h2' in globals() and 'end_time_m2' in globals():
        total_time_2,plus = break_time_func(start_time_h2, start_time_m2, end_time_h2, end_time_m2,int(comfirm_month_2),int(comfirm_day_2),int(comfirm_break_time_2))
        comfirm_nightshift_2 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v2.set(str(total_time_2) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
     
def end_number_change_m3(event):
    global end_time_m3
    global start_time_m4
    global total_time_3
    global comfirm_end_time_m_3
    global comfirm_start_time_m_4
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v4.set(endtime_m_num_combobox_3.get())
    end_time_m3 = start_time_m4 = comfirm_end_time_m_3 = comfirm_start_time_m_4 = int(endtime_m_num_combobox_3.get())

    if 'start_time_h3' in globals() and 'start_time_m3' in globals() and 'end_time_h3' in globals() and 'end_time_m3' in globals():
        total_time_3,plus = break_time_func(start_time_h3, start_time_m3, end_time_h3, end_time_m3,int(comfirm_month_3),int(comfirm_day_3),int(comfirm_break_time_3))
        comfirm_nightshift_3 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v3.set(str(total_time_3) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
        
def end_number_change_m4(event):
    global end_time_m4
    global start_time_m5
    global total_time_4
    global comfirm_end_time_m_4
    global comfirm_start_time_m_5
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v5.set(endtime_m_num_combobox_4.get())
    end_time_m4 = start_time_m5 = comfirm_end_time_m_4 = comfirm_start_time_m_5 = int(endtime_m_num_combobox_4.get())
     
    if 'start_time_h4' in globals() and 'start_time_m4' in globals() and 'end_time_h4' in globals() and 'end_time_m4' in globals():
        total_time_4,plus = break_time_func(start_time_h4, start_time_m4, end_time_h4, end_time_m4,int(comfirm_month_4),int(comfirm_day_4),int(comfirm_break_time_4))    
        comfirm_nightshift_4 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v4.set(str(total_time_4) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
        
def end_number_change_m5(event):
    global end_time_m5
    global start_time_m6
    global total_time_5
    global comfirm_end_time_m_5
    global comfirm_start_time_m_6
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v6.set(endtime_m_num_combobox_5.get())
    end_time_m5 = start_time_m6 = comfirm_end_time_m_5 = comfirm_start_time_m_6 = int(endtime_m_num_combobox_5.get())

    if 'start_time_h5' in globals() and 'start_time_m5' in globals() and 'end_time_h5' in globals() and 'end_time_m5' in globals():
        total_time_5,plus = break_time_func(start_time_h5, start_time_m5, end_time_h5, end_time_m5,int(comfirm_month_5),int(comfirm_day_5),int(comfirm_break_time_5))    
        comfirm_nightshift_5 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v5.set(str(total_time_5) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
        
def end_number_change_m6(event):
    global end_time_m6
    global start_time_m7
    global total_time_6
    global comfirm_end_time_m_6
    global comfirm_start_time_m_7
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v7.set(endtime_m_num_combobox_6.get())
    end_time_m6 = start_time_m7 = comfirm_end_time_m_6 = comfirm_start_time_m_7 = int(endtime_m_num_combobox_6.get())

    if 'start_time_h6' in globals() and 'start_time_m6' in globals() and 'end_time_h6' in globals() and 'end_time_m6' in globals():
        total_time_6,plus = break_time_func(start_time_h6, start_time_m6, end_time_h6, end_time_m6,int(comfirm_month_6),int(comfirm_day_6),int(comfirm_break_time_6))    

        comfirm_nightshift_6 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v6.set(str(total_time_6) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)

def end_number_change_m7(event):
    global end_time_m7
    global start_time_m8
    global total_time_7
    global comfirm_end_time_m_7
    global comfirm_start_time_m_8
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_v8.set(endtime_m_num_combobox_7.get())
    end_time_m7 = start_time_m8 = comfirm_end_time_m_7 = comfirm_start_time_m_8 = int(endtime_m_num_combobox_7.get())

    if 'start_time_h7' in globals() and 'start_time_m7' in globals() and 'end_time_h7' in globals() and 'end_time_m7' in globals():
        total_time_7,plus = break_time_func(start_time_h7, start_time_m7, end_time_h7, end_time_m7,int(comfirm_month_7),int(comfirm_day_7),int(comfirm_break_time_7))    

        comfirm_nightshift_7 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v7.set(str(total_time_7) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
        
def end_number_change_m8(event):
    global end_time_m8
    global total_time_8
    global comfirm_end_time_m_8
    global overwark_time
    global total_time
    global comfirm_nightshift

    time_m_num_str_m8 = endtime_m_num_combobox_8.get()
    end_time_m8  = comfirm_end_time_m_8 = int(endtime_m_num_combobox_8.get())

    if 'start_time_h8' in globals() and 'start_time_m8' in globals() and 'end_time_h8' in globals() and 'end_time_m8' in globals():
        total_time_8,plus = break_time_func(start_time_h8, start_time_m8, end_time_h8, end_time_m8,int(comfirm_month_8),int(comfirm_day_8),int(comfirm_break_time_8))
        
        comfirm_nightshift_8 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v8.set(str(total_time_8) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
def break_time_func_1(event):
    global comfirm_break_time_1
    global comfirm_nightshift
    global total_time_1
    global overwark_time
    global total_time

    comfirm_break_time_1 = break_listbox_1.get()

    if 'start_time_h1' in globals() and 'start_time_m1' in globals() and 'end_time_h1' in globals() and 'end_time_m1' in globals():
        total_time_1,plus = break_time_func(start_time_h1, start_time_m1, end_time_h1, end_time_m1,int(comfirm_month_1),int(comfirm_day_1),int(comfirm_break_time_1))
        comfirm_nightshift_1 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v1.set(str(total_time_1) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def break_time_func_2(event):
    global comfirm_break_time_2
    global comfirm_nightshift
    global total_time_2
    global overwark_time
    global total_time

    comfirm_break_time_2 = break_listbox_2.get()

    if 'start_time_h2' in globals() and 'start_time_m2' in globals() and 'end_time_h2' in globals() and 'end_time_m2' in globals():
        total_time_2,plus = break_time_func(start_time_h2, start_time_m2, end_time_h2, end_time_m2,int(comfirm_month_2),int(comfirm_day_2),int(comfirm_break_time_2))
        comfirm_nightshift_2 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v2.set(str(total_time_2) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def break_time_func_3(event):
    global comfirm_break_time_3
    global comfirm_nightshift
    global total_time_3
    global overwark_time
    global total_time

    comfirm_break_time_3 = break_listbox_3.get()

    if 'start_time_h3' in globals() and 'start_time_m3' in globals() and 'end_time_h3' in globals() and 'end_time_m3' in globals():
        total_time_3,plus = break_time_func(start_time_h3, start_time_m3, end_time_h3, end_time_m3,int(comfirm_month_3),int(comfirm_day_3),int(comfirm_break_time_3))
        comfirm_nightshift_3 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v3.set(str(total_time_3) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def break_time_func_4(event):
    global comfirm_break_time_4
    global comfirm_nightshift
    global total_time_4
    global overwark_time
    global total_time
    
    comfirm_break_time_4 = break_listbox_4.get()

    if 'start_time_h4' in globals() and 'start_time_m4' in globals() and 'end_time_h4' in globals() and 'end_time_m4' in globals():
        total_time_4,plus = break_time_func(start_time_h4, start_time_m4, end_time_h4, end_time_m4,int(comfirm_month_4),int(comfirm_day_4),int(comfirm_break_time_4))    
        comfirm_nightshift_4 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v4.set(str(total_time_4) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def break_time_func_5(event):
    global comfirm_break_time_5
    global comfirm_nightshift
    global total_time_5
    global overwark_time
    global total_time
    
    comfirm_break_time_5 = break_listbox_5.get()

    if 'start_time_h5' in globals() and 'start_time_m5' in globals() and 'end_time_h5' in globals() and 'end_time_m5' in globals():
        total_time_5,plus = break_time_func(start_time_h5, start_time_m5, end_time_h5, end_time_m5,int(comfirm_month_5),int(comfirm_day_5),int(comfirm_break_time_5))    
        comfirm_nightshift_5 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v5.set(str(total_time_5) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)
    
    Calculation()       
    
def break_time_func_6(event):
    global comfirm_break_time_6
    global comfirm_nightshift
    global total_time_6
    global overwark_time
    global total_time
    
    comfirm_break_time_6 = break_listbox_6.get()

    if 'start_time_h6' in globals() and 'start_time_m6' in globals() and 'end_time_h6' in globals() and 'end_time_m6' in globals():
        total_time_6,plus = break_time_func(start_time_h6, start_time_m6, end_time_h6, end_time_m6,int(comfirm_month_6),int(comfirm_day_6),int(comfirm_break_time_6))    
        comfirm_nightshift_6 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v6.set(str(total_time_6) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)

        Calculation() 

def break_time_func_7(event):
    global comfirm_break_time_7
    global comfirm_nightshift
    global total_time_7
    global overwark_time
    global total_time
    
    comfirm_break_time_7 = break_listbox_7.get()

    if 'start_time_h7' in globals() and 'start_time_m7' in globals() and 'end_time_h7' in globals() and 'end_time_m7' in globals():
        total_time_7,plus = break_time_func(start_time_h7, start_time_m7, end_time_h7, end_time_m7,int(comfirm_month_7),int(comfirm_day_7),int(comfirm_break_time_7))    

        comfirm_nightshift_7 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v7.set(str(total_time_7) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)   

        Calculation()

def break_time_func_8(event):
    global comfirm_break_time_8
    global comfirm_nightshift
    global total_time_8
    global overwark_time
    global total_time
    
    comfirm_break_time_8 = break_listbox_8.get()

    if 'start_time_h8' in globals() and 'start_time_m8' in globals() and 'end_time_h8' in globals() and 'end_time_m8' in globals():
        total_time_8,plus = break_time_func(start_time_h8, start_time_m8, end_time_h8, end_time_m8,int(comfirm_month_8),int(comfirm_day_8),int(comfirm_break_time_8))
        
        comfirm_nightshift_8 = nightshift(plus)

        total_time = total_time_1 + total_time_2 + total_time_3 + total_time_4 + total_time_5 + total_time_6 + total_time_7 + total_time_8
        total_label_v8.set(str(total_time_8) + " 分")
        total_time_min_v.set("合計時間 : "+ (str(total_time)) + " 分")

        comfirm_nightshift = comfirm_nightshift_1 + comfirm_nightshift_2 + comfirm_nightshift_3 + comfirm_nightshift_4 + comfirm_nightshift_5 + comfirm_nightshift_6 + comfirm_nightshift_7 + comfirm_nightshift_8
        nightshift_time_min_v.set("夜勤時間 : "+ str(comfirm_nightshift) + " 分")
            
        overwark_time = total_func(total_time)

        Calculation()

def break_time_func(start_time_h,start_time_m,end_time_h,end_time_m,month_a,day_a,break_time):
    global total_time
    global start_time
    global end_time
    global overwark_time
    global total_time
    if kanto_mode == 0:
        start_time = datetime.datetime(ny,nm,nd,start_time_h,start_time_m)
        end_time = datetime.datetime(ny,nm,nd,end_time_h,end_time_m)
        if end_time == time_23_59:
            plus = 1
        else:
            plus = 0
        delta = end_time - start_time
        total = (delta.total_seconds() / 60)
        if mode == 0:
            if break_10_10 > start_time:
                if break_15_10 <= end_time:
                    total -= 80
                elif break_13_00 <= end_time:
                    total -= 70        
                elif break_10_10 <= end_time:
                    total -= 10       
            elif break_13_00 > start_time:
                if break_15_10 <= end_time:
                    total -= 70 
                elif break_13_00 <= end_time and break_15_10 >= end_time:
                    total -= 60
            elif break_15_10 > start_time:
                if break_15_10 < end_time:
                    total -= 10
        elif mode == 1:
            total -= break_time

        total += plus
        return total,plus
    
    
    elif kanto_mode == 1:
        start_time = datetime.datetime(ny,nm,nd,start_time_h,start_time_m)
        end_time = datetime.datetime(ny,nm,nd,end_time_h,end_time_m)
        if end_time == time_23_59:
            plus = 1
        else:
            plus = 0
        delta = end_time - start_time
        total = (delta.total_seconds() / 60)
        if mode == 0:
            if break_10_15 > start_time:
                if break_15_15 <= end_time:
                    total -= 90
                elif break_13_00 <= end_time:
                    total -= 75       
                elif break_10_15 <= end_time:
                    total -= 15       
            elif break_13_00 > start_time:
                if break_15_15 <= end_time:
                    total -= 75 
                elif break_13_00 <= end_time and break_15_15 >= end_time:
                    total -= 60
            elif break_15_10 > start_time:
                if break_15_10 < end_time:
                    total -= 15
        elif mode == 1:
            total -= break_time

        total += plus
        return total,plus

def nightshift(plus):
    if  end_time > time_22_00:
        nightshift_time = end_time - time_22_00     
        nightshift_time = (nightshift_time.total_seconds() / 60)
        nightshift_time += plus
    elif end_time <= time_5_00 and start_time <= time_5_00:
        nightshift_time = end_time - start_time
        nightshift_time = (nightshift_time.total_seconds() / 60)
    elif end_time > time_5_00 and start_time < time_5_00:
        nightshift_time = time_5_00 - start_time
        nightshift_time = (nightshift_time.total_seconds() / 60) 
    else:
        nightshift_time = 0
    return nightshift_time
      
def total_func(total_time):     
    if total_time > 480.0:
        overwark_time = total_time - 480
        overwark_time_min_v.set("残業時間 : "+ (str(overwark_time)) + " 分")
    else:
        overwark_time = 0
        overwark_time_min_v.set("残業時間 : "+ (str(overwark_time)) + " 分")
    return overwark_time


from openpyxl.styles.borders import Border, Side
from tkinter.ttk import Combobox
from tkinter import messagebox
from tkinter import *
import datetime
import openpyxl
import tkinter as tk

# 日付取得
dt_now = datetime.datetime.now()
ny = now_year = dt_now.year
nm = now_month = dt_now.month
nd = now_day = dt_now.day
today = datetime.date(ny,nm,nd)

# エクセルファイル読込

loadingpass_wd = openpyxl.load_workbook('excelpath.xlsx', keep_vba=True)
loadingpass_ws = loadingpass_wd.worksheets[0]

failps1 = loadingpass_ws['B1']

failps3 = loadingpass_ws['B3']

kanto_value = loadingpass_ws['D1'].value
class_value = loadingpass_ws['D3'].value
name_value = loadingpass_ws['D4'].value

if kanto_value == None:
    kanto_mode = 0
else:
    kanto_mode = int(kanto_value)

if class_value == None:
    class_value = 0
else:
    class_mode = int(class_value)

if name_value == None:
    # name_value = 0
    pass
else:
    name_mode = int(name_value)

if class_mode == 0:
    CELL_work_name = "C"
    CELL_work_class = "N"
    CELL_work_number = "W"
    failps2 = loadingpass_ws['B6']

if class_mode == 1:
    CELL_work_name = "D"
    CELL_work_class = "O"
    CELL_work_number = "X"
    failps2 = loadingpass_ws['B7']

if class_mode == 2:
    CELL_work_name = "E"
    CELL_work_class = "P"
    CELL_work_number = "Y"
    failps2 = loadingpass_ws['B8']

if class_mode == 3:
    CELL_work_name = "F"
    CELL_work_class = "Q"
    CELL_work_number = "Z"
    failps2 = loadingpass_ws['B9']

if class_mode == 4:
    CELL_work_name = "G"
    CELL_work_class = "N"
    CELL_work_number = "W"
    failps2 = loadingpass_ws['B10']

if class_mode == 5:
    CELL_work_name = "H"
    CELL_work_class = "P"
    CELL_work_number = "Y"
    failps2 = loadingpass_ws['B11']

if class_mode == 6:
    CELL_work_name = "I"
    CELL_work_class = "R"
    CELL_work_number = "Z"
    failps2 = loadingpass_ws['B12']

if class_mode == 7:
    CELL_work_name = "J"
    CELL_work_class = "S"
    CELL_work_number = "AA"
    failps2 = loadingpass_ws['B13']

if class_mode == 8:
    CELL_work_name = "K"
    CELL_work_class = "T"
    CELL_work_number = "AB"
    failps2 = loadingpass_ws['B14']

failpass1 = str(failps1.value)
failpass2 = str(failps2.value)
failpass3 = str(failps3.value)

wb = openpyxl.load_workbook(failpass1, keep_vba=True)
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

ws = wb.worksheets[0]

break_10_10 = datetime.datetime(ny,nm,nd,10,10)
break_10_15 = datetime.datetime(ny,nm,nd,10,15)
break_13_00 = datetime.datetime(ny,nm,nd,13,00)
break_15_10 = datetime.datetime(ny,nm,nd,15,10)
break_15_15 = datetime.datetime(ny,nm,nd,15,15)
time_22_00 = datetime.datetime(ny,nm,nd,22,00)
time_23_59 = datetime.datetime(ny,nm,nd,23,59)
time_5_00 = datetime.datetime(ny,nm,nd,5,00)
time_0_00 = datetime.datetime(ny,nm,nd,0,00)

total_time_1 = 0
total_time_2 = 0
total_time_3 = 0
total_time_4 = 0
total_time_5 = 0
total_time_6 = 0
total_time_7 = 0
total_time_8 = 0 

comfirm_break_time_1 = 0
comfirm_break_time_2 = 0
comfirm_break_time_3 = 0
comfirm_break_time_4 = 0
comfirm_break_time_5 = 0
comfirm_break_time_6 = 0
comfirm_break_time_7 = 0
comfirm_break_time_8 = 0

comfirm_nightshift_1 = 0
comfirm_nightshift_2 = 0
comfirm_nightshift_3 = 0
comfirm_nightshift_4 = 0
comfirm_nightshift_5 = 0
comfirm_nightshift_6 = 0
comfirm_nightshift_7 = 0
comfirm_nightshift_8 = 0
# mode初期値
mode = 0
overwark_check_mode = 0
fix_mode = 0
All_contents = []

# class_value = loadingpass_ws['D1'].value

mat = cell_check(ws['A'], ws)
mat -= 1
mat_number = []
mat_number_name = []
for row_1 in ws["A2:A" + str(mat)]:
    for col_1 in row_1:
        mat_number.append(str(col_1.value)[:7])
        mat_number_name.append(str(col_1.value))
        mat_number_name_str = [str(n) for n in mat_number_name]
        mat_number_str = [str(n) for n in mat_number]

warkclass = cell_check(ws[CELL_work_class], ws)
warkclass -= 1
warkclass_number = []
for row_2 in ws[CELL_work_class + "2:"+ CELL_work_class + str(warkclass)]:
    for col_2 in row_2:
        warkclass_number.append(str(col_2.value))

wark = cell_check(ws[CELL_work_number], ws)
wark -= 1
wark_number = []
for row_3 in ws[CELL_work_number + "2:" + CELL_work_number + str(wark)]:
    for col_3 in row_3:
        wark_number.append(col_3.value)
        wark_number_str = [str(n) for n in wark_number]

NAME = cell_check(ws[CELL_work_name], ws)
NAME -= 1
name_number = []
for row_5 in ws[CELL_work_name + "2:"+ CELL_work_name + str(NAME)]:
    for col_5 in row_5:
        name_number.append(col_5.value)
        

# リスト作成
month_list = ["4","5","6","7","8","9","10","11","12","1","2","3"]
day_list = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31"]
time_h_list = ["8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","0","1","2","3","4","5","6","7"]
time_m_list = ["0","10","20","30","40","50"]
time_m1_list = ["0","10","20","30","40","50"]
endtime_h_list = ["8","9","10","11","12","13","14","15","16","17","18","19","20","21","22","23","0","1","2","3","4","5","6","7"]
endtime_m_list = ["0","10","20","30","40","50","59"]
break_list = ["0","10","20","30","40","50","60","70","80","90","100","110","120"]
overwark_check_list = ["0","30","60","90","120","180","210","270"]

# tkinter
version = "1.1.0"
fonts = "游ゴシック"

if kanto_mode == 0:
    app_title ='工数入力 ' + version
elif kanto_mode == 1:
    app_title ='関東工場 工数入力 ' + version

root1 = tk.Tk()
root1.title(app_title)
root1.geometry('1250x610')

name_frame = Frame(root1)
check_frame = Frame(root1,width=420, height=100, borderwidth=5, relief='ridge')
mat_frame = Frame(root1)
tag_frame = Frame(root1)
wark_frame = Frame(root1)
time_h_frame = Frame(root1)
time_m_frame = Frame(root1)
endtime_h_frame = Frame(root1)
endtime_m_frame = Frame(root1)
total_time_frame = LabelFrame(root1, text ='所要時間', width=90, height=345, borderwidth=5, relief='groove', font=(fonts,12,"bold"))
detail_frame = Frame(root1, width=100, height=40, borderwidth=5, relief='groove')
set_frame =  Frame(root1)
search_frame = Frame(root1, width=1190, height=120, borderwidth=5, relief='groove')

name_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=0, ipadx=700, ipady=40, sticky=NW)
tag_frame.grid(row=0, column=0, padx=0, pady=60, ipadx=500, ipady=0, sticky=NW)
mat_frame.grid(row=0, column=0, padx=0, pady=90, ipadx=200, ipady=500, sticky=NW)
wark_frame.grid(row=0, column=0, padx=610, pady=90, sticky=NW)
time_h_frame.grid(row=0, column=0, padx=780, pady=90, sticky=NW)
time_m_frame.grid(row=0, column=0, padx=840, pady=90, sticky=NW)
endtime_h_frame.grid(row=0, column=0, padx=920, pady=90, sticky=NW)
endtime_m_frame.grid(row=0, column=0, padx=980, pady=90, sticky=NW)
total_time_frame.grid(row=0, column=0, padx=1155, pady=67, sticky=NW)
total_time_frame.grid_propagate(0)
detail_frame.grid(row=0, column=0, padx=550, pady=415, ipadx=280, ipady=5, sticky=NW)
detail_frame.grid_propagate(0)
set_frame.grid(row=0, column=0, padx=800, pady=510, ipadx=0, sticky=NW)
search_frame.grid(row=0, column=0, padx=40, pady=410, ipadx=10, ipady=0, sticky=NW)

# --------------------------------------------------------------------------------
ward_var_1 = StringVar()
ward_var_1.set("名前選択")
ward_value_1 = tk.Label(name_frame, textvariable=ward_var_1, font=(fonts,20,"bold"))
ward_value_1.place(x=150, y=8)

name_num_v = StringVar()
name_num_combobox = Combobox(name_frame, textvariable=name_num_v, width=10, font=(fonts,15,"bold"))
name_num_combobox.grid(row=0, column=0, padx=0, pady=0, ipadx=0, sticky=W)

name_num_combobox.bind( '<<ComboboxSelected>>',func1)

fixbutton = tk.Button(name_frame, text="修正", command=cell_fix, font=(fonts,15,"bold"))
fixbutton.place(x=1050, y=5)

allclearbutton = tk.Button(name_frame, text='削除', command=allclear, bg="red", font=(fonts,12,"bold"))
allclearbutton.grid(row=0, column=4, padx=180, pady=5, ipadx=30, ipady=5, sticky=W)

any_button = tk.Button(name_frame, text='休 憩 入 力',  bg="#ffc800", command=list_state, font=(fonts,12,"bold"))
any_button.grid(row=0, column=3, padx=350, pady=5, ipadx=10, ipady=5, sticky=W)

any_label = tk.Label(name_frame, text='休憩時間入力モード',  bg="orange", font=(fonts,20,"bold"))

one_button = tk.Button(name_frame, text='1日一括入力',  bg="#ffc8ff", command=one_set, font=(fonts,12,"bold"))
one_button.grid(row=0, column=3, padx=200, pady=5, ipadx=10, ipady=5, sticky=W)
# --------------------------------------------------------------------------------
tag_label_1 = tk.Label(tag_frame, text='日付', font=(fonts,15,"bold"))
tag_label_2 = tk.Label(tag_frame, text='工番', font=(fonts,15,"bold"))
tag_label_3 = tk.Label(tag_frame, text='作業区分', font=(fonts,15,"bold"))
tag_label_4 = tk.Label(tag_frame, text='詳細',foreground='red', font=(fonts,15,"bold"))
tag_label_5 = tk.Label(tag_frame, text='作業番号', font=(fonts,15,"bold"))
tag_label_6 = tk.Label(tag_frame, text='開始時間', font=(fonts,15,"bold"))
tag_label_7 = tk.Label(tag_frame, text='終了時間', font=(fonts,15,"bold"))
tag_label_8 = tk.Label(tag_frame, text='休憩', font=(fonts,15,"bold"))

# --------------------------------------------------------------------------------
today_month_combobox_1 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_2 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_3 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_4 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_5 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_6 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_7 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)
today_month_combobox_8 = Combobox(mat_frame, width=2, font=(fonts,15,"bold"), height=10)

month_label_1 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_2 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_3 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_4 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_5 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_6 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_7 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))
month_label_8 = tk.Label(mat_frame, text='月', font=(fonts,15,"bold"))

day_combobox_v_1 = StringVar()
day_combobox_v_2 = StringVar()
day_combobox_v_3 = StringVar()
day_combobox_v_4 = StringVar()
day_combobox_v_5 = StringVar()
day_combobox_v_6 = StringVar()
day_combobox_v_7 = StringVar()
day_combobox_v_8 = StringVar()

today_day_combobox_1 = Combobox(mat_frame, textvariable = day_combobox_v_1, width=2, font=(fonts,15,"bold"), height=25)
today_day_combobox_2 = Combobox(mat_frame, textvariable = day_combobox_v_2, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_3 = Combobox(mat_frame, textvariable = day_combobox_v_3, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_4 = Combobox(mat_frame, textvariable = day_combobox_v_4, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_5 = Combobox(mat_frame, textvariable = day_combobox_v_5, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_6 = Combobox(mat_frame, textvariable = day_combobox_v_6, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_7 = Combobox(mat_frame, textvariable = day_combobox_v_7, width=2, font=(fonts,15,"bold"), height=25) 
today_day_combobox_8 = Combobox(mat_frame, textvariable = day_combobox_v_8, width=2, font=(fonts,15,"bold"), height=25) 

day_label_1 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_2 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_3 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_4 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_5 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_6 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_7 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))
day_label_8 = tk.Label(mat_frame, text='日', font=(fonts,15,"bold"))

mat_num_combobox_1 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_2 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_3 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_4 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_5 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_6 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_7 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)
mat_num_combobox_8 = Combobox(mat_frame, width=9, font=(fonts,15,"bold"), height=25)

checkbutton_var_1 = StringVar()
checkbutton_var_2 = StringVar()
checkbutton_var_3 = StringVar()
checkbutton_var_4 = StringVar()
checkbutton_var_5 = StringVar()
checkbutton_var_6 = StringVar()
checkbutton_var_7 = StringVar()
checkbutton_var_8 = StringVar()

checkbutton_var_1.set('0')
checkbutton_var_2.set('0')
checkbutton_var_3.set('0')
checkbutton_var_4.set('0')
checkbutton_var_5.set('0')
checkbutton_var_6.set('0')
checkbutton_var_7.set('0')
checkbutton_var_8.set('0')

none_lable = tk.Label(mat_frame, "" ,font=("",0,"bold"))

mat_serch_checkbutton1 = tk.Checkbutton(mat_frame, variable=checkbutton_var_1)
mat_serch_checkbutton2 = tk.Checkbutton(mat_frame, variable=checkbutton_var_2)
mat_serch_checkbutton3 = tk.Checkbutton(mat_frame, variable=checkbutton_var_3)
mat_serch_checkbutton4 = tk.Checkbutton(mat_frame, variable=checkbutton_var_4)
mat_serch_checkbutton5 = tk.Checkbutton(mat_frame, variable=checkbutton_var_5)
mat_serch_checkbutton6 = tk.Checkbutton(mat_frame, variable=checkbutton_var_6)
mat_serch_checkbutton7 = tk.Checkbutton(mat_frame, variable=checkbutton_var_7)
mat_serch_checkbutton8 = tk.Checkbutton(mat_frame, variable=checkbutton_var_8)

row_copybutton_0 = tk.Button(mat_frame, text='⤶', font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_1 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_1, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_2 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_2, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_3 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_3, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_4 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_4, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_5 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_5, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_6 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_6, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")
row_copybutton_7 = tk.Button(mat_frame, text='⤶',command=row_copybutton_func_7, font=(fonts,12,"normal"), anchor=tk.CENTER, bg="#CCFFFF")

warkclass_num_combobox_1 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_2 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_3 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_4 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_5 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_6 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_7 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)
warkclass_num_combobox_8 = Combobox(mat_frame, width=12, font=(fonts,15,"bold"), height=25)

warkclass_detail_entry_1 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_2 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_3 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_4 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_5 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_6 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_7 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
warkclass_detail_entry_8 = Entry(mat_frame, width=10, font=(fonts,15,"bold"))
# --------------------------------------------------------------------------------
wark_num_combobox_1 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_2 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_3 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_4 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_5 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_6 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_7 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
wark_num_combobox_8 = Combobox(wark_frame, width=12, font=(fonts,15,"bold"), height=25)
# --------------------------------------------------------------------------------
time_h_num_v1 = StringVar()
time_h_num_v2 = StringVar()
time_h_num_v3 = StringVar()
time_h_num_v4 = StringVar()
time_h_num_v5 = StringVar()
time_h_num_v6 = StringVar()
time_h_num_v7 = StringVar()
time_h_num_v8 = StringVar()

time_h_num_combobox_1 = Combobox(time_h_frame, textvariable=time_h_num_v1, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_2 = Combobox(time_h_frame, textvariable=time_h_num_v2, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_3 = Combobox(time_h_frame, textvariable=time_h_num_v3, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_4 = Combobox(time_h_frame, textvariable=time_h_num_v4, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_5 = Combobox(time_h_frame, textvariable=time_h_num_v5, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_6 = Combobox(time_h_frame, textvariable=time_h_num_v6, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_7 = Combobox(time_h_frame, textvariable=time_h_num_v7, width=3, font=(fonts,15,"bold"))
time_h_num_combobox_8 = Combobox(time_h_frame, textvariable=time_h_num_v8, width=3, font=(fonts,15,"bold"))
# --------------------------------------------------------------------------------
time_m_num_v1 = StringVar()
time_m_num_v2 = StringVar()
time_m_num_v3 = StringVar()
time_m_num_v4 = StringVar()
time_m_num_v5 = StringVar()
time_m_num_v6 = StringVar()
time_m_num_v7 = StringVar()
time_m_num_v8 = StringVar()
time_m_num_combobox_1 = Combobox(time_m_frame, textvariable=time_m_num_v1, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_2 = Combobox(time_m_frame, textvariable=time_m_num_v2, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_3 = Combobox(time_m_frame, textvariable=time_m_num_v3, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_4 = Combobox(time_m_frame, textvariable=time_m_num_v4, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_5 = Combobox(time_m_frame, textvariable=time_m_num_v5, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_6 = Combobox(time_m_frame, textvariable=time_m_num_v6, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_7 = Combobox(time_m_frame, textvariable=time_m_num_v7, width=3, font=(fonts,15,"bold"))
time_m_num_combobox_8 = Combobox(time_m_frame, textvariable=time_m_num_v8, width=3, font=(fonts,15,"bold"))
# --------------------------------------------------------------------------------
endtime_h_num_v1 = StringVar()
endtime_h_num_v2 = StringVar()
endtime_h_num_v3 = StringVar()
endtime_h_num_v4 = StringVar()
endtime_h_num_v5 = StringVar()
endtime_h_num_v6 = StringVar()
endtime_h_num_v7 = StringVar()
endtime_h_num_v8 = StringVar()
# --------------------------------------------------------------------------------
endtime_h_num_combobox_1 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v1, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_2 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v2, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_3 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v3, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_4 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v4, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_5 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v5, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_6 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v6, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_7 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v7, width=3, font=(fonts,15,"bold"))
endtime_h_num_combobox_8 = Combobox(endtime_h_frame, textvariable=endtime_h_num_v8, width=3, font=(fonts,15,"bold"))
# --------------------------------------------------------------------------------
endtime_m_num_v1 = StringVar()
endtime_m_num_v2 = StringVar()
endtime_m_num_v3 = StringVar()
endtime_m_num_v4 = StringVar()
endtime_m_num_v5 = StringVar()
endtime_m_num_v6 = StringVar()
endtime_m_num_v7 = StringVar()
endtime_m_num_v8 = StringVar()
# --------------------------------------------------------------------------------
endtime_m_num_combobox_1 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v1, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_2 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v2, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_3 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v3, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_4 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v4, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_5 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v5, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_6 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v6, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_7 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v7, width=3, font=(fonts,15,"bold"))
endtime_m_num_combobox_8 = Combobox(endtime_m_frame, textvariable=endtime_m_num_v8, width=3, font=(fonts,15,"bold"))

break_listbox_1 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_2 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_3 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_4 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_5 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_6 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_7 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))
break_listbox_8 = Combobox(endtime_m_frame, width=3,state="disabled", font=(fonts,15,"bold"))

row_clear_button_1 = tk.Button(endtime_m_frame, command=row1clear, bg="#FF4F02")
row_clear_button_2 = tk.Button(endtime_m_frame, command=row2clear, bg="#FF4F02")
row_clear_button_3 = tk.Button(endtime_m_frame, command=row3clear, bg="#FF4F02")
row_clear_button_4 = tk.Button(endtime_m_frame, command=row4clear, bg="#FF4F02")
row_clear_button_5 = tk.Button(endtime_m_frame, command=row5clear, bg="#FF4F02")
row_clear_button_6 = tk.Button(endtime_m_frame, command=row6clear, bg="#FF4F02")
row_clear_button_7 = tk.Button(endtime_m_frame, command=row7clear, bg="#FF4F02")
row_clear_button_8 = tk.Button(endtime_m_frame, command=row8clear, bg="#FF4F02")
# ----------------------------------------------------------------------------------------
total_label_v1 = StringVar()
total_label_v2 = StringVar()
total_label_v3 = StringVar()
total_label_v4 = StringVar()
total_label_v5 = StringVar()
total_label_v6 = StringVar()
total_label_v7 = StringVar()
total_label_v8 = StringVar()
# --------------------------------------------------------------------------------
total_label_1 = tk.Label(total_time_frame, textvariable=total_label_v1, font=(fonts,13,"bold"))
total_label_2 = tk.Label(total_time_frame, textvariable=total_label_v2, font=(fonts,13,"bold"))
total_label_3 = tk.Label(total_time_frame, textvariable=total_label_v3, font=(fonts,13,"bold"))
total_label_4 = tk.Label(total_time_frame, textvariable=total_label_v4, font=(fonts,13,"bold"))
total_label_5 = tk.Label(total_time_frame, textvariable=total_label_v5, font=(fonts,13,"bold"))
total_label_6 = tk.Label(total_time_frame, textvariable=total_label_v6, font=(fonts,13,"bold"))
total_label_7 = tk.Label(total_time_frame, textvariable=total_label_v7, font=(fonts,13,"bold"))
total_label_8 = tk.Label(total_time_frame, textvariable=total_label_v8, font=(fonts,13,"bold"))
# --------------------------------------------------------------------------------
total_time_min_v = StringVar()
total_time_min = tk.Label(detail_frame, textvariable=total_time_min_v, font=(fonts,15,"bold"))
total_time_min_v.set("合計時間 : "+ "0" + " 分")

overwark_time_min_v = StringVar()
overwark_time_min = tk.Label(detail_frame, textvariable=overwark_time_min_v, font=(fonts,15,"bold"))
overwark_time_min_v.set("残業時間 : "+ "0" + " 分")

nightshift_time_min_v = StringVar()
nightshift_time_min = tk.Label(detail_frame, textvariable=nightshift_time_min_v, font=(fonts,15,"bold"))
nightshift_time_min_v.set("夜勤時間 : "+ "0" + " 分")
# --------------------------------------------------------------------------------
set_button = tk.Button(set_frame, text='登　録', font=(fonts,20,"bold"), bg="#7FFF00", command=set1)
# --------------------------------------------------------------------------------
searchbox_label = tk.Label(search_frame, text='工番検索', font=(fonts,15,"bold"))
searchbox_label.place(x=1000, y=5)

searchbox_entry = Entry(search_frame, width=30, font=(fonts,15,"bold"))
searchbox_entry.grid(row=1, column=0, padx=10, pady=1)

searchbox_button = tk.Button(search_frame, text='検　索', font=(fonts,12,"bold"), command=search)
searchbox_button.grid(row=1, column=1, padx=0, pady=1)

searchbox_listbox = Listbox(search_frame,height=5, width=30, font=(fonts,15,"bold"))
searchbox_listbox.grid(row=2, column=0, padx=14, pady=3)

searchbox_button_clear = tk.Button(search_frame, text='挿　入', font=(fonts,12,"bold"), command=search_insert)
searchbox_button_clear.grid(row=2, column=1, padx=0, pady=1)
# # --------------------------------------------------------------------------------
name_num_combobox['values'] = name_number
try:
    name_num_combobox.current(name_mode)
    comfirm_name = name_num_combobox.get()
    ward_var_1.set(comfirm_name)
except NameError:
    pass
# --------------------------------------------------------------------------------
month = month_list.index(str(today.month))
today_month_combobox_1['values'] = month_list
today_month_combobox_1.current(month)
today_month_combobox_2['values'] = month_list
today_month_combobox_2.current(month)
today_month_combobox_3['values'] = month_list
today_month_combobox_3.current(month)
today_month_combobox_4['values'] = month_list
today_month_combobox_4.current(month)
today_month_combobox_5['values'] = month_list
today_month_combobox_5.current(month)
today_month_combobox_6['values'] = month_list
today_month_combobox_6.current(month)
today_month_combobox_7['values'] = month_list
today_month_combobox_7.current(month)
today_month_combobox_8['values'] = month_list
today_month_combobox_8.current(month)
comfirm_month_1 = int(today_month_combobox_1.get())
comfirm_month_2 = int(today_month_combobox_2.get())
comfirm_month_3 = int(today_month_combobox_3.get())
comfirm_month_4 = int(today_month_combobox_4.get())
comfirm_month_5 = int(today_month_combobox_5.get())
comfirm_month_6 = int(today_month_combobox_6.get())
comfirm_month_7 = int(today_month_combobox_7.get())
comfirm_month_8 = int(today_month_combobox_8.get())


day = day_list.index(str(today.day))
today_day_combobox_1['values'] = day_list
today_day_combobox_2['values'] = day_list
today_day_combobox_3['values'] = day_list
today_day_combobox_4['values'] = day_list
today_day_combobox_5['values'] = day_list
today_day_combobox_6['values'] = day_list
today_day_combobox_7['values'] = day_list
today_day_combobox_8['values'] = day_list

today_day_combobox_1.current(day)
today_day_combobox_2.current(day)
today_day_combobox_3.current(day)
today_day_combobox_4.current(day)
today_day_combobox_5.current(day)
today_day_combobox_6.current(day)
today_day_combobox_7.current(day)
today_day_combobox_8.current(day)

comfirm_day_1 = int(today_day_combobox_1.get())
comfirm_day_2 = int(today_day_combobox_2.get())
comfirm_day_3 = int(today_day_combobox_3.get())
comfirm_day_4 = int(today_day_combobox_4.get())
comfirm_day_5 = int(today_day_combobox_5.get())
comfirm_day_6 = int(today_day_combobox_6.get())
comfirm_day_7 = int(today_day_combobox_7.get())
comfirm_day_8 = int(today_day_combobox_8.get())

mat_num_combobox_1['values'] = mat_number
mat_num_combobox_2['values'] = mat_number
mat_num_combobox_3['values'] = mat_number
mat_num_combobox_4['values'] = mat_number
mat_num_combobox_5['values'] = mat_number
mat_num_combobox_6['values'] = mat_number
mat_num_combobox_7['values'] = mat_number
mat_num_combobox_8['values'] = mat_number

warkclass_num_combobox_1['values'] = warkclass_number 
warkclass_num_combobox_2['values'] = warkclass_number
warkclass_num_combobox_3['values'] = warkclass_number
warkclass_num_combobox_4['values'] = warkclass_number
warkclass_num_combobox_5['values'] = warkclass_number
warkclass_num_combobox_6['values'] = warkclass_number
warkclass_num_combobox_7['values'] = warkclass_number
warkclass_num_combobox_8['values'] = warkclass_number

wark_num_combobox_1['values'] = wark_number
wark_num_combobox_2['values'] = wark_number
wark_num_combobox_3['values'] = wark_number
wark_num_combobox_4['values'] = wark_number
wark_num_combobox_5['values'] = wark_number
wark_num_combobox_6['values'] = wark_number
wark_num_combobox_7['values'] = wark_number
wark_num_combobox_8['values'] = wark_number

time_h_num_combobox_1['values'] = time_h_list
time_h_num_combobox_2['values'] = time_h_list
time_h_num_combobox_3['values'] = time_h_list
time_h_num_combobox_4['values'] = time_h_list
time_h_num_combobox_5['values'] = time_h_list
time_h_num_combobox_6['values'] = time_h_list
time_h_num_combobox_7['values'] = time_h_list
time_h_num_combobox_8['values'] = time_h_list

time_m_num_combobox_1['values'] = time_m1_list
time_m_num_combobox_2['values'] = time_m_list
time_m_num_combobox_3['values'] = time_m_list
time_m_num_combobox_4['values'] = time_m_list
time_m_num_combobox_5['values'] = time_m_list
time_m_num_combobox_6['values'] = time_m_list
time_m_num_combobox_7['values'] = time_m_list
time_m_num_combobox_8['values'] = time_m_list
if kanto_mode == 0:
    time_h_num_combobox_1.current(0)
    time_m_num_combobox_1.current(1)
elif kanto_mode == 1:
    time_h_num_combobox_1.current(0)
    time_m_num_combobox_1.current(0)
comfirm_start_time_h_1 = start_time_h1=int(time_h_num_combobox_1.get())
comfirm_start_time_m_1 = start_time_m1=int(time_m_num_combobox_1.get())

endtime_h_num_combobox_1['values'] = endtime_h_list
endtime_h_num_combobox_2['values'] = endtime_h_list
endtime_h_num_combobox_3['values'] = endtime_h_list
endtime_h_num_combobox_4['values'] = endtime_h_list
endtime_h_num_combobox_5['values'] = endtime_h_list
endtime_h_num_combobox_6['values'] = endtime_h_list
endtime_h_num_combobox_7['values'] = endtime_h_list
endtime_h_num_combobox_8['values'] = endtime_h_list

endtime_m_num_combobox_1['values'] = endtime_m_list
endtime_m_num_combobox_2['values'] = endtime_m_list
endtime_m_num_combobox_3['values'] = endtime_m_list
endtime_m_num_combobox_4['values'] = endtime_m_list
endtime_m_num_combobox_5['values'] = endtime_m_list
endtime_m_num_combobox_6['values'] = endtime_m_list
endtime_m_num_combobox_7['values'] = endtime_m_list
endtime_m_num_combobox_8['values'] = endtime_m_list

break_listbox_1['values'] = break_list
break_listbox_2['values'] = break_list
break_listbox_3['values'] = break_list
break_listbox_4['values'] = break_list
break_listbox_5['values'] = break_list
break_listbox_6['values'] = break_list
break_listbox_7['values'] = break_list
break_listbox_8['values'] = break_list

# overwark_check_combobox['values'] = overwark_check_list

# --------------------------------------------------------------------------------
name_num_combobox.bind( '<<ComboboxSelected>>')
# --------------------------------------------------------------------------------

today_month_combobox_1.bind( '<<ComboboxSelected>>',today_month_1)
today_month_combobox_2.bind( '<<ComboboxSelected>>',today_month_2)
today_month_combobox_3.bind( '<<ComboboxSelected>>',today_month_3)
today_month_combobox_4.bind( '<<ComboboxSelected>>',today_month_4)
today_month_combobox_5.bind( '<<ComboboxSelected>>',today_month_5)
today_month_combobox_6.bind( '<<ComboboxSelected>>',today_month_6)
today_month_combobox_7.bind( '<<ComboboxSelected>>',today_month_7)
today_month_combobox_8.bind( '<<ComboboxSelected>>',today_month_8)

today_day_combobox_1.bind( '<<ComboboxSelected>>',today_day_1)
today_day_combobox_2.bind( '<<ComboboxSelected>>',today_day_2)
today_day_combobox_3.bind( '<<ComboboxSelected>>',today_day_3)
today_day_combobox_4.bind( '<<ComboboxSelected>>',today_day_4)
today_day_combobox_5.bind( '<<ComboboxSelected>>',today_day_5)
today_day_combobox_6.bind( '<<ComboboxSelected>>',today_day_6)
today_day_combobox_7.bind( '<<ComboboxSelected>>',today_day_7)
today_day_combobox_8.bind( '<<ComboboxSelected>>',today_day_8)

mat_num_combobox_1.bind( '<<ComboboxSelected>>', mat_num_1)
mat_num_combobox_2.bind( '<<ComboboxSelected>>', mat_num_2)
mat_num_combobox_3.bind( '<<ComboboxSelected>>', mat_num_3)
mat_num_combobox_4.bind( '<<ComboboxSelected>>', mat_num_4)
mat_num_combobox_5.bind( '<<ComboboxSelected>>', mat_num_5)
mat_num_combobox_6.bind( '<<ComboboxSelected>>', mat_num_6)
mat_num_combobox_7.bind( '<<ComboboxSelected>>', mat_num_7)
mat_num_combobox_8.bind( '<<ComboboxSelected>>', mat_num_8)

wark_num_combobox_1.bind( '<<ComboboxSelected>>', wark_num_1)
wark_num_combobox_2.bind( '<<ComboboxSelected>>', wark_num_2)
wark_num_combobox_3.bind( '<<ComboboxSelected>>', wark_num_3)
wark_num_combobox_4.bind( '<<ComboboxSelected>>', wark_num_4)
wark_num_combobox_5.bind( '<<ComboboxSelected>>', wark_num_5)
wark_num_combobox_6.bind( '<<ComboboxSelected>>', wark_num_6)
wark_num_combobox_7.bind( '<<ComboboxSelected>>', wark_num_7)
wark_num_combobox_8.bind( '<<ComboboxSelected>>', wark_num_8)

warkclass_num_combobox_1.bind( '<<ComboboxSelected>>', warkclass_num_1)
warkclass_num_combobox_2.bind( '<<ComboboxSelected>>', warkclass_num_2)
warkclass_num_combobox_3.bind( '<<ComboboxSelected>>', warkclass_num_3)
warkclass_num_combobox_4.bind( '<<ComboboxSelected>>', warkclass_num_4)
warkclass_num_combobox_5.bind( '<<ComboboxSelected>>', warkclass_num_5)
warkclass_num_combobox_6.bind( '<<ComboboxSelected>>', warkclass_num_6)
warkclass_num_combobox_7.bind( '<<ComboboxSelected>>', warkclass_num_7)
warkclass_num_combobox_8.bind( '<<ComboboxSelected>>', warkclass_num_8)

warkclass_detail_entry_1.bind( '<Motion>', warkclass_detail_get_1)
warkclass_detail_entry_2.bind( '<Motion>', warkclass_detail_get_2)
warkclass_detail_entry_3.bind( '<Motion>', warkclass_detail_get_3)
warkclass_detail_entry_4.bind( '<Motion>', warkclass_detail_get_4)
warkclass_detail_entry_5.bind( '<Motion>', warkclass_detail_get_5)
warkclass_detail_entry_6.bind( '<Motion>', warkclass_detail_get_6)
warkclass_detail_entry_7.bind( '<Motion>', warkclass_detail_get_7)
warkclass_detail_entry_8.bind( '<Motion>', warkclass_detail_get_8)

time_h_num_combobox_1.bind( '<<ComboboxSelected>>', start_number_change_h1)
time_h_num_combobox_2.bind( '<<ComboboxSelected>>', start_number_change_h2)
time_h_num_combobox_3.bind( '<<ComboboxSelected>>', start_number_change_h3)
time_h_num_combobox_4.bind( '<<ComboboxSelected>>', start_number_change_h4)
time_h_num_combobox_5.bind( '<<ComboboxSelected>>', start_number_change_h5)
time_h_num_combobox_6.bind( '<<ComboboxSelected>>', start_number_change_h6)
time_h_num_combobox_7.bind( '<<ComboboxSelected>>', start_number_change_h7)
time_h_num_combobox_8.bind( '<<ComboboxSelected>>', start_number_change_h8)

time_m_num_combobox_1.bind( '<<ComboboxSelected>>', start_number_change_m1)
time_m_num_combobox_2.bind( '<<ComboboxSelected>>', start_number_change_m2)
time_m_num_combobox_3.bind( '<<ComboboxSelected>>', start_number_change_m3)
time_m_num_combobox_4.bind( '<<ComboboxSelected>>', start_number_change_m4)
time_m_num_combobox_5.bind( '<<ComboboxSelected>>', start_number_change_m5)
time_m_num_combobox_6.bind( '<<ComboboxSelected>>', start_number_change_m6)
time_m_num_combobox_7.bind( '<<ComboboxSelected>>', start_number_change_m7)
time_m_num_combobox_8.bind( '<<ComboboxSelected>>', start_number_change_m8)


endtime_h_num_combobox_1.bind( '<<ComboboxSelected>>', end_number_change_h1)
endtime_h_num_combobox_2.bind( '<<ComboboxSelected>>', end_number_change_h2)
endtime_h_num_combobox_3.bind( '<<ComboboxSelected>>', end_number_change_h3)
endtime_h_num_combobox_4.bind( '<<ComboboxSelected>>', end_number_change_h4)
endtime_h_num_combobox_5.bind( '<<ComboboxSelected>>', end_number_change_h5)
endtime_h_num_combobox_6.bind( '<<ComboboxSelected>>', end_number_change_h6)
endtime_h_num_combobox_7.bind( '<<ComboboxSelected>>', end_number_change_h7)
endtime_h_num_combobox_8.bind( '<<ComboboxSelected>>', end_number_change_h8)


endtime_m_num_combobox_1.bind( '<<ComboboxSelected>>', end_number_change_m1)
endtime_m_num_combobox_2.bind( '<<ComboboxSelected>>', end_number_change_m2)
endtime_m_num_combobox_3.bind( '<<ComboboxSelected>>', end_number_change_m3)
endtime_m_num_combobox_4.bind( '<<ComboboxSelected>>', end_number_change_m4)
endtime_m_num_combobox_5.bind( '<<ComboboxSelected>>', end_number_change_m5)
endtime_m_num_combobox_6.bind( '<<ComboboxSelected>>', end_number_change_m6)
endtime_m_num_combobox_7.bind( '<<ComboboxSelected>>', end_number_change_m7)
endtime_m_num_combobox_8.bind( '<<ComboboxSelected>>', end_number_change_m8)


break_listbox_1.bind( '<<ComboboxSelected>>', break_time_func_1)
break_listbox_2.bind( '<<ComboboxSelected>>', break_time_func_2)
break_listbox_3.bind( '<<ComboboxSelected>>', break_time_func_3)
break_listbox_4.bind( '<<ComboboxSelected>>', break_time_func_4)
break_listbox_5.bind( '<<ComboboxSelected>>', break_time_func_5)
break_listbox_6.bind( '<<ComboboxSelected>>', break_time_func_6)
break_listbox_7.bind( '<<ComboboxSelected>>', break_time_func_7)
break_listbox_8.bind( '<<ComboboxSelected>>', break_time_func_8)


# --------------------------------------------------------------------------------

# --------------------------------------------------------------------------------
tag_label_1.grid(row=0, column=0, padx=50)
tag_label_2.place(x=205, y=1)
tag_label_3.place(x=364, y=1)
tag_label_4.place(x=533, y=1)
tag_label_5.place(x=655, y=1)
tag_label_6.place(x=791, y=1)
tag_label_7.place(x=933, y=1)
tag_label_8.place(x=1040, y=1)

today_month_combobox_1.grid(row=0, column=0, padx=10, pady=4)
today_month_combobox_2.grid(row=1, column=0, padx=10, pady=4)
today_month_combobox_3.grid(row=2, column=0, padx=10, pady=4)
today_month_combobox_4.grid(row=3, column=0, padx=10, pady=4)
today_month_combobox_5.grid(row=4, column=0, padx=10, pady=4)
today_month_combobox_6.grid(row=5, column=0, padx=10, pady=4)
today_month_combobox_7.grid(row=6, column=0, padx=10, pady=4)
today_month_combobox_8.grid(row=7, column=0, padx=10, pady=4)

month_label_1.place(x=54.5, y=5)
month_label_2.place(x=54.5, y=45)
month_label_3.place(x=54.5, y=85)
month_label_4.place(x=54.5, y=125)
month_label_5.place(x=54.5, y=165)
month_label_6.place(x=54.5, y=205)
month_label_7.place(x=54.5, y=245)
month_label_8.place(x=54.5, y=285)

today_day_combobox_1.grid(row=0, column=2, padx=25, pady=4)
today_day_combobox_2.grid(row=1, column=2, padx=25, pady=4)
today_day_combobox_3.grid(row=2, column=2, padx=25, pady=4)
today_day_combobox_4.grid(row=3, column=2, padx=25, pady=4)
today_day_combobox_5.grid(row=4, column=2, padx=25, pady=4)
today_day_combobox_6.grid(row=5, column=2, padx=25, pady=4)
today_day_combobox_7.grid(row=6, column=2, padx=25, pady=4)
today_day_combobox_8.grid(row=7, column=2, padx=25, pady=4)

day_label_1.place(x=135, y=5)
day_label_2.place(x=135, y=45)
day_label_3.place(x=135, y=85)
day_label_4.place(x=135, y=125)
day_label_5.place(x=135, y=165)
day_label_6.place(x=135, y=205)
day_label_7.place(x=135, y=245)
day_label_8.place(x=135, y=285)

mat_num_combobox_1.grid(row=0, column=4, padx=6, pady=4)
mat_num_combobox_2.grid(row=1, column=4, padx=6, pady=4)
mat_num_combobox_3.grid(row=2, column=4, padx=6, pady=4)
mat_num_combobox_4.grid(row=3, column=4, padx=6, pady=4)
mat_num_combobox_5.grid(row=4, column=4, padx=6, pady=4)
mat_num_combobox_6.grid(row=5, column=4, padx=6, pady=4)
mat_num_combobox_7.grid(row=6, column=4, padx=6, pady=4)
mat_num_combobox_8.grid(row=7, column=4, padx=6, pady=4)

none_lable.grid(row=7, column=5, padx=15, pady=4)

row_copybutton_0['state'] = "disabled"
row_copybutton_0.place(x=298,y=19 ,width = 20, height = 20)
row_copybutton_1.place(x=298,y=59 ,width = 20, height = 20)
row_copybutton_2.place(x=298,y=99 ,width = 20, height = 20)
row_copybutton_3.place(x=298,y=139 ,width = 20, height = 20)
row_copybutton_4.place(x=298,y=179 ,width = 20, height = 20)
row_copybutton_5.place(x=298,y=219 ,width = 20, height = 20)
row_copybutton_6.place(x=298,y=259 ,width = 20, height = 20)
row_copybutton_7.place(x=298,y=299 ,width = 20, height = 20)

mat_serch_checkbutton1.place(x=298, y=0)
mat_serch_checkbutton2.place(x=298, y=40)
mat_serch_checkbutton3.place(x=298, y=80)
mat_serch_checkbutton4.place(x=298, y=120)
mat_serch_checkbutton5.place(x=298, y=160)
mat_serch_checkbutton6.place(x=298, y=200)
mat_serch_checkbutton7.place(x=298, y=240)
mat_serch_checkbutton8.place(x=298, y=280)

warkclass_num_combobox_1.grid(row=0, column=6, padx=0, pady=4)
warkclass_num_combobox_2.grid(row=1, column=6, padx=0, pady=4)
warkclass_num_combobox_3.grid(row=2, column=6, padx=0, pady=4)
warkclass_num_combobox_4.grid(row=3, column=6, padx=0, pady=4)
warkclass_num_combobox_5.grid(row=4, column=6, padx=0, pady=4)
warkclass_num_combobox_6.grid(row=5, column=6, padx=0, pady=4)
warkclass_num_combobox_7.grid(row=6, column=6, padx=0, pady=4)
warkclass_num_combobox_8.grid(row=7, column=6, padx=0, pady=4)

warkclass_detail_entry_1.grid(row=0, column=7, padx=14, pady=4)
warkclass_detail_entry_2.grid(row=1, column=7, padx=14, pady=4)
warkclass_detail_entry_3.grid(row=2, column=7, padx=14, pady=4)
warkclass_detail_entry_4.grid(row=3, column=7, padx=14, pady=4)
warkclass_detail_entry_5.grid(row=4, column=7, padx=14, pady=4)
warkclass_detail_entry_6.grid(row=5, column=7, padx=14, pady=4)
warkclass_detail_entry_7.grid(row=6, column=7, padx=14, pady=4)
warkclass_detail_entry_8.grid(row=7, column=7, padx=14, pady=4)

wark_num_combobox_1.grid(padx=10, pady=4)
wark_num_combobox_2.grid(padx=10, pady=4)
wark_num_combobox_3.grid(padx=10, pady=4)
wark_num_combobox_4.grid(padx=10, pady=4)
wark_num_combobox_5.grid(padx=10, pady=4)
wark_num_combobox_6.grid(padx=10, pady=4)
wark_num_combobox_7.grid(padx=10, pady=4)
wark_num_combobox_8.grid(padx=10, pady=4)

time_h_num_combobox_1.grid(pady=4)
time_h_num_combobox_2.grid(pady=4)
time_h_num_combobox_3.grid(pady=4)
time_h_num_combobox_4.grid(pady=4)
time_h_num_combobox_5.grid(pady=4)
time_h_num_combobox_6.grid(pady=4)
time_h_num_combobox_7.grid(pady=4)
time_h_num_combobox_8.grid(pady=4)


time_m_num_combobox_1.grid(pady=4)
time_m_num_combobox_2.grid(pady=4)
time_m_num_combobox_3.grid(pady=4)
time_m_num_combobox_4.grid(pady=4)
time_m_num_combobox_5.grid(pady=4)
time_m_num_combobox_6.grid(pady=4)
time_m_num_combobox_7.grid(pady=4)
time_m_num_combobox_8.grid(pady=4)

endtime_h_num_combobox_1.grid(pady=4)
endtime_h_num_combobox_2.grid(pady=4)
endtime_h_num_combobox_3.grid(pady=4)
endtime_h_num_combobox_4.grid(pady=4)
endtime_h_num_combobox_5.grid(pady=4)
endtime_h_num_combobox_6.grid(pady=4)
endtime_h_num_combobox_7.grid(pady=4)
endtime_h_num_combobox_8.grid(pady=4)

endtime_m_num_combobox_1.grid(pady=4)
endtime_m_num_combobox_2.grid(pady=4)
endtime_m_num_combobox_3.grid(pady=4)
endtime_m_num_combobox_4.grid(pady=4)
endtime_m_num_combobox_5.grid(pady=4)
endtime_m_num_combobox_6.grid(pady=4)
endtime_m_num_combobox_7.grid(pady=4)
endtime_m_num_combobox_8.grid(pady=4)

break_listbox_1.grid(row=0, column=1,padx=5, pady=4)
break_listbox_2.grid(row=1, column=1,padx=5, pady=4)
break_listbox_3.grid(row=2, column=1,padx=5, pady=4)
break_listbox_4.grid(row=3, column=1,padx=5, pady=4)
break_listbox_5.grid(row=4, column=1,padx=5, pady=4)
break_listbox_6.grid(row=5, column=1,padx=5, pady=4)
break_listbox_7.grid(row=6, column=1,padx=5, pady=4)
break_listbox_8.grid(row=7, column=1,padx=5, pady=4)

row_clear_button_1.grid(row=0, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_2.grid(row=1, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_3.grid(row=2, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_4.grid(row=3, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_5.grid(row=4, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_6.grid(row=5, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_7.grid(row=6, column=2, padx=1, pady=4, ipadx=17)
row_clear_button_8.grid(row=7, column=2, padx=1, pady=4, ipadx=17)

total_label_1.grid(pady=6, padx=1)
total_label_2.grid(pady=6, padx=1)
total_label_3.grid(pady=6, padx=1)
total_label_4.grid(pady=6, padx=1)
total_label_5.grid(pady=6, padx=1)
total_label_6.grid(pady=6, padx=1)
total_label_7.grid(pady=6, padx=1)
total_label_8.grid(pady=6, padx=1)


total_time_min.place(x=10,y=5)
overwark_time_min.place(x=230,y=5)
nightshift_time_min.place(x=450,y=5)
set_button.pack(padx=100, ipadx=100, ipady=10)





root1.resizable(width=False, height=False)
root1.mainloop()





